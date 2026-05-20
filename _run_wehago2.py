"""Fresh Wehago automation - full pipeline"""
import asyncio
import sys
import os
import time
import openpyxl
from playwright.async_api import async_playwright

if sys.platform == "win32":
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.detach(), encoding="utf-8")
    sys.stderr = io.TextIOWrapper(sys.stderr.detach(), encoding="utf-8")

CDP_URL = "http://localhost:9222"
SAVE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), "results"))
COMPANY_NAME = "근린커피 상암"


def log(msg):
    print(msg, flush=True)


async def dismiss_dialogs(page):
    for _ in range(20):
        closed = await page.evaluate("""() => {
            const selectors = ['._isDialog', '.LUX_basic_dialog'];
            let target = null;
            for (const sel of selectors) {
                for (const d of document.querySelectorAll(sel)) {
                    if (d.style.display === 'none') continue;
                    target = d; break;
                }
                if (target) break;
            }
            if (!target) return null;
            const allBtns = target.querySelectorAll('button, a');
            for (const btn of allBtns) {
                if (btn.textContent.trim() === '닫기') { btn.click(); return '닫기'; }
            }
            const luxBtns = target.querySelectorAll('button.WSC_LUXButton');
            for (const btn of luxBtns) {
                if (!btn.textContent.trim()) { btn.click(); return 'X'; }
            }
            for (const btn of allBtns) {
                if (btn.textContent.trim() === '확인') { btn.click(); return '확인'; }
            }
            for (const btn of allBtns) {
                if (btn.textContent.trim() === '취소') { btn.click(); return '취소'; }
            }
            return 'stuck';
        }""")
        if not closed:
            return
        log(f"  팝업 닫음 ({closed})")
        await asyncio.sleep(0.5)


async def click_dialog_button(page, button_text):
    await page.evaluate("""(btnText) => {
        const selectors = ['._isDialog', '.LUX_basic_dialog'];
        let target = null;
        for (const sel of selectors) {
            for (const d of document.querySelectorAll(sel)) {
                if (d.style.display !== 'none') { target = d; break; }
            }
            if (target) break;
        }
        if (!target) return;
        const btns = target.querySelectorAll('button, a');
        for (const b of btns) {
            if (b.textContent.trim().includes(btnText)) { b.click(); return; }
        }
    }""", button_text)
    await asyncio.sleep(1)
    log(f"  모달 버튼 클릭: {button_text}")


async def _click_modal_text(page, text_fragment, action):
    for _ in range(20):
        result = await page.evaluate("""(args) => {
            const all = document.querySelectorAll('*');
            for (const el of all) {
                if (!el.textContent.includes(args.fragment)) continue;
                const btns = el.querySelectorAll('button');
                for (const btn of btns) {
                    if (btn.textContent.trim() === args.action && btn.offsetWidth > 0) {
                        btn.click();
                        return args.action;
                    }
                }
            }
            return null;
        }""", {"fragment": text_fragment, "action": action})
        if result:
            return True
        await asyncio.sleep(0.5)
    return False


async def main():
    os.makedirs(SAVE_DIR, exist_ok=True)

    async with async_playwright() as p:
        log("[1] Chrome CDP 연결...")
        browser = await p.chromium.connect_over_cdp(CDP_URL)
        context = browser.contexts[0]
        page = context.pages[0]
        log(f"  현재: {await page.title()} | {page.url}")

        await dismiss_dialogs(page)

        # 현재 페이지 상태 감지
        on_smarta_salary = "SWSA0101" in page.url
        on_smarta_main = "smarta.wehago.com" in page.url and "SWSA0101" not in page.url
        on_wehago_main = "wehago.com" in page.url and "smarta" not in page.url

        # [2] 수임처 급여 페이지 이동
        if on_smarta_salary:
            log("  이미 급여자료입력 페이지에 있음. Step 2-4 스킵.")
        elif on_smarta_main:
            log("  SmartA 메인에 있음. Step 2 스킵.")
        else:
            log(f'[2] 수임처 "{COMPANY_NAME}" 급여 페이지 이동...')
            log(f'[2] 수임처 "{COMPANY_NAME}" 급여 페이지 이동...')
            await page.evaluate("""() => {
                window.__capturedUrl = null;
                window.__origOpen = window.open;
                window.open = function(url) {
                    window.__capturedUrl = url;
                    return null;
                };
            }""")

            clicked = await page.evaluate("""(companyName) => {
                const allDivs = document.querySelectorAll('[id^="company_"]');
                for (const div of allDivs) {
                    const nameEl = div.querySelector('a');
                    if (nameEl && nameEl.textContent.trim() === companyName) {
                        let card = div;
                        for (let i = 0; i < 3; i++) card = card.parentElement;
                        const buttons = card.querySelectorAll('button.btn_quick');
                        for (const btn of buttons) {
                            if (btn.querySelector('span')?.textContent.trim() === '급여') {
                                btn.click();
                                return true;
                            }
                        }
                    }
                }
                return false;
            }""", COMPANY_NAME)

            if not clicked:
                log(f'  "{COMPANY_NAME}" 급여 버튼을 찾지 못함.')
                await page.screenshot(path=os.path.join(SAVE_DIR, "debug_step2.png"))
                return

            await asyncio.sleep(1)
            salary_url = await page.evaluate("() => window.__capturedUrl")
            await page.evaluate("() => { window.open = window.__origOpen; }")
            log(f"  SmartA URL: {salary_url}")

            if not salary_url:
                log("  URL 캡처 실패")
                return

            await page.goto(salary_url, wait_until="domcontentloaded", timeout=30000)
            await asyncio.sleep(3)
            log(f"  이동 완료: {await page.title()}")
            await dismiss_dialogs(page)

        # [3] 급여자료입력 메뉴
        if not on_smarta_salary:
            log("[3] 급여자료입력 메뉴 이동...")
            menu_link = page.locator("a#SWSA0101")
            if await menu_link.count() > 0:
                await menu_link.first.click()
                log("  SWSA0101 클릭")
            else:
                log("  SWSA0101 없음")
            await asyncio.sleep(5)
            log(f"  URL: {page.url}")

            # 간이세액 모달 닫기
            log("[3-1] 간이세액 안내 모달 닫기...")
            await page.evaluate("""() => {
                const all = document.querySelectorAll('*');
                for (const el of all) {
                    const cs = window.getComputedStyle(el);
                    if (cs.position !== 'fixed' || cs.display === 'none' ||
                        parseInt(cs.zIndex) <= 100 || el.offsetWidth <= 100) continue;
                    if (!el.textContent.includes('간이세액')) continue;
                    const btns = el.querySelectorAll('button.WSC_LUXButton');
                    for (const btn of btns) {
                        if (!btn.textContent.trim() && btn.offsetWidth > 0) { btn.click(); return; }
                    }
                }
            }""")
            await asyncio.sleep(1)
            await dismiss_dialogs(page)
        else:
            log("[3] 이미 급여자료입력 페이지 - 스킵")

        # [4] 구분 드롭다운
        log("[4] 구분 드롭다운 -> 급여+상여...")
        await page.evaluate("""() => {
            const dd = document.querySelectorAll('.LS_ngh_select2')[0];
            if (dd) dd.querySelector('.LSbutton').click();
        }""")
        await asyncio.sleep(1)
        await page.evaluate("""(text) => {
            const items = document.querySelectorAll('.LSselectResult li');
            for (const li of items) {
                if (li.textContent.includes(text)) {
                    li.querySelector('a').click();
                    return true;
                }
            }
            return false;
        }""", "급여+상여")
        await asyncio.sleep(1)
        value = await page.evaluate("""() => {
            const dd = document.querySelectorAll('.LS_ngh_select2')[0];
            return dd ? dd.querySelector('.fakeinput').textContent.trim() : '';
        }""")
        log(f"  선택됨: {value}")

        # 복사후 재계산 (조건부)
        await asyncio.sleep(1)
        has_modal = await page.evaluate("""() => {
            const selectors = ['._isDialog', '.LUX_basic_dialog'];
            for (const sel of selectors) {
                for (const d of document.querySelectorAll(sel)) {
                    if (d.style.display !== 'none') return true;
                }
            }
            return false;
        }""")
        if has_modal:
            log("[4-1] 복사후 재계산...")
            await click_dialog_button(page, "복사후 재계산")
            await asyncio.sleep(1)
            await click_dialog_button(page, "취소")

        # [5] 엑셀 다운로드
        log("[5] 엑셀 다운로드...")
        await page.evaluate("""() => {
            const btn = document.querySelector('#collect');
            if (btn) btn.click();
        }""")
        await asyncio.sleep(1)

        download_future = asyncio.Future()
        def on_download(d):
            if not download_future.done():
                log(f"  다운로드 감지: {d.suggested_filename}")
                download_future.set_result(d)
        page.on("download", on_download)

        await page.evaluate("""(text) => {
            const menu = document.querySelector('.sao_head_menu');
            if (!menu) return false;
            const items = menu.querySelectorAll('li');
            for (const li of items) {
                if (li.textContent.includes(text)) {
                    const a = li.querySelector('a');
                    if (a) { a.click(); return true; }
                    li.click();
                    return true;
                }
            }
            return false;
        }""", "엑셀 내려받기")

        try:
            download = await asyncio.wait_for(download_future, timeout=15)
        except asyncio.TimeoutError:
            log("  다운로드 이벤트 없음.")
            return

        fname = download.suggested_filename
        download_path = os.path.join(SAVE_DIR, fname)
        await download.save_as(download_path)

        file_size = os.path.getsize(download_path) if os.path.exists(download_path) else 0
        log(f"  다운로드 완료: {download_path} ({file_size} bytes)")

        if file_size == 0:
            log("  파일이 0 bytes. 중단.")
            return

        # [6] 업로드 양식 변환
        log("[6] 업로드 양식 변환...")
        wb_src = openpyxl.load_workbook(download_path)
        ws_src = wb_src["Sheet1"]

        headers = []
        for c in range(1, ws_src.max_column + 1):
            h2 = ws_src.cell(2, c).value
            h1 = ws_src.cell(1, c).value
            if h2 and str(h2).strip():
                headers.append(str(h2).strip())
            elif h1 and str(h1).strip():
                headers.append(str(h1).strip())
            else:
                headers.append(None)

        TEXT_COLS = {"사원코드", "사원명", "부서", "직급", "직종"}
        wb_new = openpyxl.Workbook()
        ws_new = wb_new.active
        ws_new.title = "Sheet1"

        for i, header in enumerate(headers, 1):
            ws_new.cell(1, i).value = header

        new_row = 2
        for r in range(3, ws_src.max_row + 1):
            first_val = ws_src.cell(r, 1).value
            if not first_val or first_val == "합계":
                continue
            for c in range(1, ws_src.max_column + 1):
                val = ws_src.cell(r, c).value
                header = headers[c - 1]
                if header == "사원코드" and isinstance(val, str):
                    try:
                        val = str(int(val)).zfill(4)
                    except (ValueError, TypeError):
                        pass
                if val is None:
                    val = "" if header in TEXT_COLS else 0
                ws_new.cell(new_row, c).value = val
            new_row += 1

        upload_path = os.path.join(SAVE_DIR, "wehago_upload.xlsx")
        wb_new.save(upload_path)
        log(f"  변환 완료: {upload_path}")

        # [7] 엑셀 업로드
        log("[7] 엑셀 업로드...")
        await page.evaluate("""() => {
            const btn = document.querySelector('#collect');
            if (btn) btn.click();
        }""")
        await asyncio.sleep(1)

        log("  엑셀 불러오기 클릭...")
        async with page.expect_file_chooser(timeout=15000) as fc_info:
            await page.evaluate("""(text) => {
                const menu = document.querySelector('.sao_head_menu');
                if (!menu) return false;
                const items = menu.querySelectorAll('li');
                for (const li of items) {
                    if (li.textContent.includes(text)) {
                        const a = li.querySelector('a');
                        if (a) { a.click(); return true; }
                        li.click();
                        return true;
                    }
                }
                return false;
            }""", "엑셀 불러오기")

        file_chooser = await fc_info.value
        log(f"  파일 선택: {os.path.basename(upload_path)}")
        await file_chooser.set_files(upload_path)
        await asyncio.sleep(3)

        # 헤더 행 선택
        log("  1) 헤더 행 선택...")
        clicked = await page.evaluate("""() => {
            const tables = document.querySelectorAll('table');
            for (const table of tables) {
                if (table.offsetParent === null) continue;
                const trs = table.querySelectorAll('tr');
                if (trs.length > 2) {
                    const th = trs[1].querySelector('th');
                    if (th && th.textContent.trim() === '1') {
                        th.click();
                        return true;
                    }
                }
            }
            return false;
        }""")
        log(f"  행1 클릭: {clicked}")
        await asyncio.sleep(1)

        # 엑셀제목설정
        log("  2) 엑셀제목설정...")
        await page.evaluate("""() => {
            const btns = document.querySelectorAll('button.WSC_LUXButton');
            for (const btn of btns) {
                if (btn.textContent.trim() === '② 엑셀제목설정') {
                    btn.click();
                    return;
                }
            }
        }""")
        await asyncio.sleep(2)
        await _click_modal_text(page, "엑셀제목", "확인")
        await asyncio.sleep(2)

        # 확인
        log("  확인 버튼 클릭...")
        await page.evaluate("""() => {
            const selectors = ['._isDialog', '.LUX_basic_dialog'];
            for (const sel of selectors) {
                for (const dialog of document.querySelectorAll(sel)) {
                    if (dialog.style.display === 'none' || dialog.offsetParent === null) continue;
                    const btns = dialog.querySelectorAll('button.WSC_LUXButton');
                    for (const btn of btns) {
                        if (btn.textContent.trim() === '확인') {
                            btn.click();
                            return;
                        }
                    }
                }
            }
        }""")
        await asyncio.sleep(5)

        # 후속 모달들
        log("  후속 모달 1 -> #confirm 확인...")
        await page.evaluate("""() => {
            const btn = document.querySelector('#confirm');
            if (btn) btn.click();
        }""")
        await asyncio.sleep(3)

        log("  후속 모달 2 -> '연결되지 않은 사원' 확인...")
        await _click_modal_text(page, "연결되지 않은 사원", "확인")
        await asyncio.sleep(3)

        log("  후속 모달 3 -> '삭제후 업로드' 취소...")
        await _click_modal_text(page, "삭제후 업로드", "취소")
        await asyncio.sleep(3)

        log("  후속 모달 4 -> '변환이 취소' 확인...")
        await _click_modal_text(page, "변환이 취소", "확인")
        await asyncio.sleep(2)

        await page.screenshot(path=os.path.join(SAVE_DIR, "wehago_final.png"))
        log("\n완료!")


if __name__ == "__main__":
    asyncio.run(main())
