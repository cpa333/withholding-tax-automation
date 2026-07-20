"""E2E 검증: WEHAGO 급여자료입력 원천데이터 반영 — 단독/병렬/혼용 레이아웃 모두.

실제 포맷 파일(NHIS PDF·NPS 통합엑셀·고용보험 xls·WEHAGO 2행헤더 엑셀)을
합성해 디스크에 배치하고, 내가 수정한 전체 체인을 구동한다:
  _locate_raw_data  →  read_* 파서  →  convert_for_upload(+apply_raw_data)  →  업로드 엑셀

각 레이아웃×사원×보험 셀 값이 기대치와 일치하는지 검증.
P1(게이트 확장), P2(EI 0보존), P0(병렬 경로 폴백) 모두 시나리오에 포함.

실행: python -m scripts.e2e_swsa_merge_verify
      (또는) python scripts/e2e_swsa_merge_verify.py

의존성(수동 검증용 — requirements 에 불포함):
  pip install xlwt reportlab   (고용보험 .xls / NHIS PDF 합성용)
한글 폰트: C:\\Windows\\Fonts\\malgun.ttf (reportlab 기본 폰트는 한글 미지원 →
          폰트 미등록 시 pdfplumber 가 '건강'/'요양' 을 깨먹어 NHIS 파서가 빈 dict 반환).
프로덕션 NHIS/고용보험 파서는 실제 포털 PDF/xls (한글 폰트 임베드) 를 읽으므로 본 합성 이슈와 무관.
"""
import os
import sys
import shutil
import tempfile

# repo root 를 path 에 (스크립트 위치: <repo>/scripts/)
REPO = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
sys.path.insert(0, REPO)


# ════════════════════════════════════════════════════════════════════
# 0. 시나리오 데이터 (4명 사원 — 각 보험의 모든 분기를 커버)
# ════════════════════════════════════════════════════════════════════
# 홍길동  : NHIS(건강/요양) + NPS(member-only) + EI(support-only)
# 김철수  : NHIS + NPS(member+retro, 정산컬럼없음→국민연금에 누적) + EI(collect-only)
# 박보험  : EI support==collect → adjustment=0 → 고용보험 재계산값 보존(P2). WEHAGO 초기값 15000 유지
# 최국고  : NPS(member+govt) → 국민연금에서 govt 차감. EI 없음.
EMPLOYEES = ["홍길동", "김철수", "박보험", "최국고"]

# NHIS (산출_건강, 산출_요양) — 정산/연말/이자=0
NHIS = {
    "홍길동": (100000, 50000),
    "김철수": (80000, 40000),
}
# NPS 통합엑셀 (member, retro, govt) — govt 는 이미 분할된 본인몫
NPS = {
    "홍길동": dict(member=200000, retro=0, govt=0),
    "김철수": dict(member=150000, retro=10000, govt=0),
    "최국고": dict(member=100000, retro=0, govt=20000),
}
# 고용보험 (support, collect)
EI = {
    "홍길동": dict(support=30000, collect=0),
    "김철수": dict(support=0, collect=20000),
    "박보험": dict(support=40000, collect=40000),  # → adjustment 0 → 셀 보존
}
# WEHAGO 다운로드 엑셀의 초기 고용보험값 (재계산 결과 시뮬레이션).
# EI 조정분은 이 값을 **기준**으로 가감된다 (덮어쓰기 아님).
WEHAGO_EI_INITIAL = {"홍길동": 40000, "김철수": 18000, "박보험": 15000, "최국고": 0}

# 기대치 (최종 업로드 엑셀 셀 값)
EXPECTED = {
    # name: (건강보험, 장기요양보험료, 국민연금, 고용보험)
    "홍길동": (100000, 50000, 200000, 10000),    # EI: 재계산 40000 - |30000|
    "김철수": (80000, 40000, 160000, 38000),     # NPS: member+retro(정산col없음→누적)=160000, EI: 재계산 18000 + |20000|
    "박보험": (0, 0, 0, 15000),                  # EI adjustment=0 → 재계산값 보존(P2)
    "최국고": (0, 0, 80000, 0),                  # NPS: member-govt=80000, EI 없음 → 재계산값 0 유지
}


# ════════════════════════════════════════════════════════════════════
# 1. 파일 합성기
# ════════════════════════════════════════════════════════════════════
def build_nhis_pdf(path):
    """reportlab 16컬럼 표 → pdfplumber 추출 round-trip.
    컬럼 idx: 0=성명 4=구분 5=산출 7=사유 9=정산 12=연말 15=이자.
    한글 폰트(Malgun Gothic) 등록 필수 — 기본 Helvetica 는 한글 미지원.
    """
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
    from reportlab.lib import colors
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    font_path = r"C:\Windows\Fonts\malgun.ttf"
    if os.path.exists(font_path):
        pdfmetrics.registerFont(TTFont("Malgun", font_path))
        font_name = "Malgun"
    else:
        font_name = "Helvetica"  # 폰트 없으면 한글 깨짐(파서가 무시) — 폴백

    header = ["성명", "c1", "c2", "c3", "구분", "산출보험료", "c6", "사유",
              "c8", "정산금액", "c10", "c11", "연말정산", "c13", "c14", "환급금이자"]
    rows = [header]
    for name, (h, y) in NHIS.items():
        # 건강 행 (idx0=이름)
        rows.append([name, "", "", "", "건강", str(h), "", "", "",
                     "0", "", "", "0", "", "", "0"])
        # 요양 행 (idx0=None → 직전 이름 사용)
        rows.append(["", "", "", "", "요양", str(y), "", "", "",
                     "0", "", "", "0", "", "", "0"])
    doc = SimpleDocTemplate(path, pagesize=A4)
    t = Table(rows)
    t.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
        ("FONTNAME", (0, 0), (-1, -1), font_name),
    ]))
    doc.build([t])


def build_nps_integrated_xlsx(path):
    """openpyxl — 통합엑셀 (R1 헤더). col3=성명 col10=member col16=retro col24=govt."""
    from openpyxl import Workbook
    HEADERS = [
        "고지년월", "사업장관리번호", "성명", "주민등록번호", "취득일", "상실일",
        "당월분_기준소득월액(원)", "당월분_월보험료(원)",
        "당월분_(사용자부담금)(원)", "당월분_(본인기여금(원)",
        "소급분_변동사유", "소급분_해당기간", "소급분_월수",
        "소급분_전월이전보험료(원)", "소급분_(사용자부담금)(원)",
        "소급분_(본인기여금)(원)", "이자분(0원)",
        "총부담금계_(사용자부담금)(원)", "총부담금계_(본인기여금)(원)",
        "자격변동 신고사항", "자격변동 신고사항_대상자",
        "국고지원금액(원)", "국고지원금액_사용자부담금(원)",
        "국고지원금액_본인기여금(원)", "취득월 납부여부",
    ]
    wb = Workbook(); ws = wb.active
    for c, val in enumerate(HEADERS, start=1):
        ws.cell(1, c).value = val
    r = 2
    for name, d in NPS.items():
        row = [None] * 25
        row[2] = name  # col3
        if d["member"]:
            row[9] = d["member"]    # col10
        if d["retro"]:
            row[15] = d["retro"]    # col16
        if d["govt"]:
            row[23] = d["govt"]     # col24
        for c, val in enumerate(row, start=1):
            ws.cell(r, c).value = val
        r += 1
    wb.save(path); wb.close()


def build_employment_xls(path):
    """xlwt — Page1(지원금: col2=이름, col12=금액), Page2(환수금: col1=이름, col7=금액).
    데이터는 행7부터 (파서 range(7, nrows))."""
    import xlwt
    wb = xlwt.Workbook()
    sh1 = wb.add_sheet("Page 1")
    for r in range(7):
        sh1.write(r, 0, f"meta{r}")
    sh1.write(6, 2, "근로자명"); sh1.write(6, 12, "실업급여지원금(근로자)")
    rr = 7
    for name, d in EI.items():
        if d["support"]:
            sh1.write(rr, 2, name); sh1.write(rr, 12, d["support"]); rr += 1
    sh2 = wb.add_sheet("Page 2")
    for r in range(7):
        sh2.write(r, 0, f"meta{r}")
    sh2.write(6, 1, "근로자명"); sh2.write(6, 7, "실업급여환수금(근로자)")
    rr = 7
    for name, d in EI.items():
        if d["collect"]:
            sh2.write(rr, 1, name); sh2.write(rr, 7, d["collect"]); rr += 1
    wb.save(path)


def build_wehago_download_xlsx(path):
    """openpyxl — WEHAGO 2행헤더 다운로드 양식 (공임나라 샘플과 동일 구조, 19컬럼).
    보험 셀은 0(또는 박보험 고용보험=15000). 합계행 포함."""
    from openpyxl import Workbook
    row1 = ["사원코드", "사원명", "부서", "직급", "직종", "수당", None, None,
            "공제", None, None, None, None, None, None, None, None, None, "차인지급액"]
    row2 = [None, None, None, None, None, "기본급", "상여", "지급액계",
            "국민연금", "건강보험", "고용보험", "장기요양보험료", "소득세",
            "지방소득세", "학자금상환액", "건강보험료정산", "장기요양보험료정산",
            "공제액계", None]
    wb = Workbook(); ws = wb.active; ws.title = "Sheet1"
    for c, v in enumerate(row1, start=1):
        ws.cell(1, c).value = v
    for c, v in enumerate(row2, start=1):
        ws.cell(2, c).value = v
    r = 3
    for i, name in enumerate(EMPLOYEES, start=1):
        vals = [str(i), name, "본부", "사원", "일반",        # 1-5
                2000000, 0, 2000000,                          # 6-8 기본급/상여/지급액계
                0, 0, WEHAGO_EI_INITIAL[name], 0,             # 9-12 국민연금/건강보험/고용보험/장기요양
                50000, 5000, 0, 0, 0, 0, 1925000]             # 13-19
        for c, v in enumerate(vals, start=1):
            ws.cell(r, c).value = v
        r += 1
    # 합계행 (convert_for_upload 이 스킵)
    ws.cell(r, 1).value = "합계"
    wb.save(path); wb.close()


# ════════════════════════════════════════════════════════════════════
# 2. 레이아웃 배치
# ════════════════════════════════════════════════════════════════════
CLIENT = "E2E회사"
PERIOD = "202607"

def place_files(root, layout):
    """layout 별로 합성 파일 배치. root 아래에 period/client/... 구조 생성."""
    client_dir_standalone = os.path.join(root, f"__client__", CLIENT)
    os.makedirs(client_dir_standalone, exist_ok=True)
    # 파일은 한 번만 합성해서 각 위치로 복사
    tmp = tempfile.mkdtemp()
    pdf = os.path.join(tmp, "가입자고지내역서_건강_20260701.pdf")
    nps = os.path.join(tmp, "결정내역통보서_202607.xlsx")
    ei = os.path.join(tmp, "고용보험료지원금정보_202607.xls")
    build_nhis_pdf(pdf); build_nps_integrated_xlsx(nps); build_employment_xls(ei)

    def put(dst_dir, src):
        os.makedirs(dst_dir, exist_ok=True)
        shutil.copy2(src, os.path.join(dst_dir, os.path.basename(src)))

    if layout == "standalone":
        put(os.path.join(root, f"국민건강보험_{PERIOD}", CLIENT), pdf)
        put(os.path.join(root, f"국민연금_{PERIOD}", CLIENT), nps)
        put(os.path.join(root, f"고용보험_{PERIOD}", CLIENT), ei)
    elif layout == "parallel":
        base = os.path.join(root, f"공단EDI_{PERIOD}", CLIENT)
        put(os.path.join(base, "국민건강보험"), pdf)
        put(os.path.join(base, "국민연금"), nps)
        put(os.path.join(base, "고용보험"), ei)
    elif layout == "mixed":
        # NHIS 단독, NPS·EI 병렬
        put(os.path.join(root, f"국민건강보험_{PERIOD}", CLIENT), pdf)
        base = os.path.join(root, f"공단EDI_{PERIOD}", CLIENT)
        put(os.path.join(base, "국민연금"), nps)
        put(os.path.join(base, "고용보험"), ei)
    shutil.rmtree(tmp, ignore_errors=True)


# ════════════════════════════════════════════════════════════════════
# 3. 전체 체인 구동 + 검증
# ════════════════════════════════════════════════════════════════════
def run_layout(layout):
    """한 레이아웃에서 전체 체인 실행. (발견키, 파싱결과 요약, 셀검증 결과) 반환."""
    import src.utils.save_path as sp
    import src.workflows.wehago_swsa as wsw
    from src.workflows.wehago_swsa import WehagoSwsaWorkflow
    from src.utils.raw_data_reader import (
        read_nhis_pdf, read_nps_integrated_excel, read_employment_xls,
    )
    from src.automation.wehago._swsa_excel import convert_for_upload

    root = tempfile.mkdtemp(prefix=f"e2e_{layout}_")
    # get_desktop_path → root.  (wsw 모듈이 import 한 바인딩을 패치)
    wsw.get_desktop_path = lambda: root

    place_files(root, layout)

    # --- _locate_raw_data (P0 핵심) ---
    raw = WehagoSwsaWorkflow._locate_raw_data(CLIENT, 2026, 7)
    found_keys = sorted(k for k, v in (raw or {}).items() if v) if raw else []

    assert raw is not None, f"[{layout}] _locate_raw_data → None (파일 못 찾음)"
    for key in ("nhis_pdf", "nps_integrated", "ei_xls"):
        assert raw.get(key), f"[{layout}] {key} 미발견"

    # --- 파서 ---
    nhis_data = read_nhis_pdf(raw["nhis_pdf"])
    nps_member, nps_retro, nps_govt = read_nps_integrated_excel(raw["nps_integrated"])
    ei_support, ei_collect = read_employment_xls(raw["ei_xls"])

    parse_summary = {
        "nhis": {n: (d.산출_건강보험료, d.산출_요양보험료) for n, d in nhis_data.items()},
        "nps_member": dict(nps_member), "nps_retro": dict(nps_retro), "nps_govt": dict(nps_govt),
        "ei_support": dict(ei_support), "ei_collect": dict(ei_collect),
    }

    # --- convert_for_upload + apply_raw_data ---
    wehago_path = os.path.join(root, "wehago_download.xlsx")
    build_wehago_download_xlsx(wehago_path)
    upload_path = convert_for_upload(
        wehago_path,
        nhis_data=nhis_data,
        nps_member_data=nps_member, nps_retro_data=nps_retro, nps_govt_data=nps_govt,
        ei_support_data=ei_support, ei_collect_data=ei_collect,
    )

    # --- 업로드 엑셀에서 셀 값 읽기 (헤더명으로 컬럼 인덱스) ---
    from openpyxl import load_workbook
    wb = load_workbook(upload_path); ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    colidx = {h: i for i, h in enumerate(headers)}
    need = ["사원명", "건강보험", "장기요양보험료", "국민연금", "고용보험"]
    for n in need:
        assert n in colidx, f"[{layout}] 업로드 헤더에 '{n}' 없음: {headers}"

    actual = {}
    for r in range(2, ws.max_row + 1):
        name = ws.cell(r, colidx["사원명"] + 1).value
        if not name or str(name).strip() == "합계":
            continue
        name = str(name).strip()
        actual[name] = tuple(
            ws.cell(r, colidx[col] + 1).value or 0
            for col in ["건강보험", "장기요양보험료", "국민연금", "고용보험"]
        )
    wb.close()

    # --- 기대치 비교 ---
    mismatches = []
    for name in EMPLOYEES:
        got = actual.get(name)
        exp = EXPECTED[name]
        if got != exp:
            mismatches.append(f"  {name}: 기대 {exp} / 실제 {got}")

    shutil.rmtree(root, ignore_errors=True)
    return found_keys, parse_summary, mismatches


# ════════════════════════════════════════════════════════════════════
# 4. main
# ════════════════════════════════════════════════════════════════════
def main():
    # wehago log 를 stdout 으로 (진단 로깅 P3 확인용)
    import src.automation.wehago._common as common
    def _log(msg, *a, **k):
        print(f"    [log] {msg}")
    common.log = _log

    print("=" * 72)
    print("E2E: WEHAGO 급여자료입력 원천데이터 반영 — 단독/병렬/혼용")
    print("=" * 72)

    all_ok = True
    for layout in ("standalone", "parallel", "mixed"):
        print(f"\n▼ 레이아웃: {layout}")
        try:
            found_keys, parse_summary, mismatches = run_layout(layout)
        except AssertionError as e:
            print(f"  ✗ FAIL: {e}")
            all_ok = False
            continue
        except Exception as e:
            import traceback
            print(f"  ✗ ERROR: {e}")
            traceback.print_exc()
            all_ok = False
            continue

        print(f"  _locate_raw_data 발견: {found_keys}")
        print(f"  파싱: NHIS={parse_summary['nhis']}")
        print(f"        NPS member={parse_summary['nps_member']} retro={parse_summary['nps_retro']} govt={parse_summary['nps_govt']}")
        print(f"        EI support={parse_summary['ei_support']} collect={parse_summary['ei_collect']}")
        if mismatches:
            print(f"  ✗ 셀 값 불일치:")
            for m in mismatches:
                print(m)
            all_ok = False
        else:
            print(f"  ✓ 4명×4컬럼(건강/요양/연금/고용) 모두 기대치 일치")

    print("\n" + "=" * 72)
    print("RESULT:", "ALL PASS ✓" if all_ok else "FAIL ✗")
    print("=" * 72)
    return 0 if all_ok else 1


if __name__ == "__main__":
    sys.exit(main())
