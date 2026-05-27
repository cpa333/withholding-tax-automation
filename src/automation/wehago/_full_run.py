"""WEHAGO 전체 자동화 — 수정 코드 적용 (dry_run: 취소)"""
import asyncio
import sys
import os
import openpyxl

if sys.platform == "win32":
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.detach(), encoding='utf-8')
    sys.stderr = io.TextIOWrapper(sys.stderr.detach(), encoding='utf-8')

# PROJECT_ROOT to sys.path for src.* imports
PROJECT_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
sys.path.insert(0, PROJECT_ROOT)

from playwright.async_api import async_playwright
from src.utils.chrome_cdp import CDP_URL

COMPANY_NAME = "근린커피 상암"
SAVE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", "results"))


def log(msg):
    print(msg, flush=True)


async def dismiss_dialogs(page):
    for _ in range(20):
        closed = await page.evaluate("""() => {
            const selectors = ['._isDialog', '.LUX_basic_dialog'];
            let target = null;
            for (const sel of selectors) {
                for (const d of document.querySelectorAll(sel)) {
                    if (d.style.display !== 'none' && d.offsetParent !== null) { target = d; break; }
                }
                if (target) break;
            }
            if (!target) return null;
            const allBtns = target.querySelectorAll('button, a');
            for (const btn of allBtns) { if (btn.textContent.trim() === '닫기') { btn.click(); return '닫기'; } }
            for (const btn of allBtns) { if (btn.textContent.trim() === '확인') { btn.click(); return '확인'; } }
            for (const btn of allBtns) { if (btn.textContent.trim() === '취소') { btn.click(); return '취소'; } }
            const luxBtns = target.querySelectorAll('button.WSC_LUXButton');
            for (const btn of luxBtns) { if (!btn.textContent.trim()) { btn.click(); return 'X'; } }
            return 'stuck';
        }""")
        if not closed:
            break
        log(f"  팝업 닫음 ({closed})")
        await asyncio.sleep(0.5)


async def click_modal_button(page, text_fragment, action):
    """특정 텍스트가 포함된 모달에서 action(확인/취소) 버튼 클릭"""
    for _ in range(20):
        result = await page.evaluate("""(args) => {
            const all = document.querySelectorAll('*');
            for (const el of all) {
                if (!el.textContent.includes(args.fragment)) continue;
                const btns = el.querySelectorAll('button');
                for (const btn of btns) {
                    if (btn.textContent.trim() === args.action && btn.offsetWidth > 0) {
                        btn.click();
                        return args.action;
                    }
                }
            }
            return null;
        }""", {"fragment": text_fragment, "action": action})
        if result:
            log(f"  [{text_fragment[:20]}...] → {result}")
            return True
        await asyncio.sleep(0.5)
    return False


async def main():
    async with async_playwright() as p:
        browser = await p.chromium.connect_over_cdp(CDP_URL)
        context = browser.contexts[0]

        from src.utils.stealth import stealth_all_pages, register_auto_stealth
        await stealth_all_pages(context)
        register_auto_stealth(context)

        page = context.pages[0]

        # ===== [1] WEHAGO 메인 =====
        log("[1] WEHAGO 메인 이동...")
        await page.goto("https://www.wehago.com/#/main", wait_until="domcontentloaded", timeout=15000)
        await asyncio.sleep(3)
        has_login = await page.locator("#company_").count() > 0 or await page.locator("text=나의 수임처").count() > 0
        if not has_login:
            log("❌ 로그인 필요")
            return
        log("  ✅ 로그인 확인\n")

        # ===== [2] 수임처 급여 =====
        log(f"[2] '{COMPANY_NAME}' 급여 이동...")
        await page.evaluate("""() => {
            window.__capturedUrl = null;
            window.__origOpen = window.open;
            window.open = function(url) { window.__capturedUrl = url; return null; };
        }""")
        clicked = await page.evaluate("""(cn) => {
            const allDivs = document.querySelectorAll('[id^="company_"]');
            for (const div of allDivs) {
                const nameEl = div.querySelector('a');
                if (nameEl && nameEl.textContent.trim() === cn) {
                    let card = div;
                    for (let i = 0; i < 3; i++) card = card.parentElement;
                    const buttons = card.querySelectorAll('button.btn_quick');
                    for (const btn of buttons) {
                        if (btn.querySelector('span')?.textContent.trim() === '급여') { btn.click(); return true; }
                    }
                }
            }
            return false;
        }""", COMPANY_NAME)
        if not clicked:
            log(f"  ❌ '{COMPANY_NAME}' 급여 버튼 못 찾음")
            return
        await asyncio.sleep(1)
        url = await page.evaluate("() => window.__capturedUrl")
        await page.evaluate("() => { window.open = window.__origOpen; }")
        if not url:
            log("  ❌ SmartA URL 없음")
            return
        await page.goto(url, wait_until="domcontentloaded", timeout=30000)
        await asyncio.sleep(3)
        log(f"  ✅ {await page.title()}\n")

        # ===== [3] 급여자료입력 =====
        log("[3] 급여자료입력 이동...")
        await page.evaluate("""() => {
            const link = document.querySelector('a#SWSA0101.text_link');
            if (link) link.click();
        }""")
        await asyncio.sleep(3)

        # 간이세액 개정 안내 모달 닫기 (X 버튼)
        log("[3-1] 간이세액 안내 모달 닫기...")
        await page.evaluate("""() => {
            const all = document.querySelectorAll('*');
            for (const el of all) {
                const cs = window.getComputedStyle(el);
                if (cs.position !== 'fixed' || cs.display === 'none' ||
                    parseInt(cs.zIndex) <= 100 || el.offsetWidth <= 100) continue;
                if (!el.textContent.includes('간이세액')) continue;
                const btns = el.querySelectorAll('button.WSC_LUXButton');
                for (const btn of btns) {
                    if (!btn.textContent.trim() && btn.offsetWidth > 0) { btn.click(); return; }
                }
            }
        }""")
        await asyncio.sleep(1)
        await dismiss_dialogs(page)
        log(f"  ✅ {await page.title()}\n")

        # ===== [4] 구분: 급여+상여 =====
        log("[4] 급여+상여 선택...")
        await page.evaluate("""(idx) => {
            const dd = document.querySelectorAll('.LS_ngh_select2')[idx];
            if (dd) dd.querySelector('.LSbutton').click();
        }""", 0)
        await asyncio.sleep(1)
        await page.evaluate("""(text) => {
            const items = document.querySelectorAll('.LSselectResult li');
            for (const li of items) {
                if (li.textContent.includes(text)) { li.querySelector('a').click(); return; }
            }
        }""", "급여+상여")
        await asyncio.sleep(1)
        log("  ✅ 급여+상여\n")

        # ===== [5-6] 복사후 재계산 모달 =====
        await asyncio.sleep(1)
        has_modal = await page.evaluate("""() => {
            const selectors = ['._isDialog', '.LUX_basic_dialog'];
            for (const sel of selectors) {
                for (const d of document.querySelectorAll(sel)) {
                    if (d.style.display !== 'none' && d.offsetParent !== null) return true;
                }
            }
            return false;
        }""")
        if has_modal:
            log("[5] 복사후 재계산 클릭...")
            await page.evaluate("""() => {
                const selectors = ['._isDialog', '.LUX_basic_dialog'];
                for (const sel of selectors) {
                    for (const d of document.querySelectorAll(sel)) {
                        if (d.style.display === 'none' || d.offsetParent === null) continue;
                        const btns = d.querySelectorAll('button, a');
                        for (const btn of btns) {
                            if (btn.textContent.trim() === '복사후 재계산') { btn.click(); return; }
                        }
                    }
                }
            }""")
            await asyncio.sleep(1)
            log("[6] 취소 클릭...")
            await page.evaluate("""() => {
                const selectors = ['._isDialog', '.LUX_basic_dialog'];
                for (const sel of selectors) {
                    for (const d of document.querySelectorAll(sel)) {
                        if (d.style.display === 'none' || d.offsetParent === null) continue;
                        const btns = d.querySelectorAll('button, a');
                        for (const btn of btns) {
                            if (btn.textContent.trim() === '취소') { btn.click(); return; }
                        }
                    }
                }
            }""")
        else:
            log("[5-6] 모달 없음 - 스킵")

        # ===== [7] 엑셀 다운로드 =====
        log("\n[7] 엑셀 다운로드...")
        await page.evaluate("""() => { const btn = document.querySelector('#collect'); if (btn) btn.click(); }""")
        await asyncio.sleep(2)

        download_future = asyncio.Future()
        def on_download(d):
            if not download_future.done():
                log(f"  다운로드: {d.suggested_filename}")
                download_future.set_result(d)
        page.on("download", on_download)

        await page.evaluate("""() => {
            const menu = document.querySelector('.sao_head_menu');
            if (!menu) return;
            const items = menu.querySelectorAll('li');
            for (const li of items) {
                if (li.textContent.includes('엑셀 내려받기')) {
                    const a = li.querySelector('a');
                    if (a) { a.click(); return; }
                }
            }
        }""")

        try:
            download = await asyncio.wait_for(download_future, timeout=15)
            os.makedirs(SAVE_DIR, exist_ok=True)
            save_path = os.path.join(SAVE_DIR, download.suggested_filename)
            await download.save_as(save_path)
            log(f"  ✅ 저장: {save_path}\n")
        except asyncio.TimeoutError:
            log("  ❌ 다운로드 시간 초과")
            return

        # ===== [8] 업로드 양식 변환 =====
        log("[8] 업로드 양식 변환...")
        wb_src = openpyxl.load_workbook(save_path)
        ws_src = wb_src["Sheet1"]
        headers = []
        for c in range(1, ws_src.max_column + 1):
            h2 = ws_src.cell(2, c).value
            h1 = ws_src.cell(1, c).value
            headers.append(str(h2).strip() if h2 and str(h2).strip() else (str(h1).strip() if h1 and str(h1).strip() else None))

        TEXT_COLS = {"사원코드", "사원명", "부서", "직급", "직종"}
        wb_new = openpyxl.Workbook()
        ws_new = wb_new.active
        ws_new.title = "Sheet1"
        for i, header in enumerate(headers, 1):
            ws_new.cell(1, i).value = header
        new_row = 2
        for r in range(3, ws_src.max_row + 1):
            first_val = ws_src.cell(r, 1).value
            if not first_val or first_val == "합계":
                continue
            for c in range(1, ws_src.max_column + 1):
                val = ws_src.cell(r, c).value
                header = headers[c - 1]
                if header == "사원코드" and isinstance(val, str):
                    try:
                        val = str(int(val)).zfill(4)
                    except (ValueError, TypeError):
                        pass
                if val is None:
                    val = "" if header in TEXT_COLS else 0
                ws_new.cell(new_row, c).value = val
            new_row += 1

        base, ext = os.path.splitext(save_path)
        upload_path = f"{base}_업로드{ext}"
        wb_new.save(upload_path)
        log(f"  ✅ 변환: {upload_path}\n")

        # ===== [9] 엑셀 업로드 =====
        log("[9] 엑셀 업로드...")

        # 드롭다운 열기
        await page.evaluate("""() => { const btn = document.querySelector('#collect'); if (btn) btn.click(); }""")
        await asyncio.sleep(2)

        # 엑셀 불러오기
        log("  엑셀 불러오기...")
        try:
            async with page.expect_file_chooser(timeout=15000) as fc_info:
                await page.evaluate("""() => {
                    const menu = document.querySelector('.sao_head_menu');
                    if (!menu) return;
                    const items = menu.querySelectorAll('li');
                    for (const li of items) {
                        if (li.textContent.includes('엑셀 불러오기')) {
                            const a = li.querySelector('a');
                            if (a) { a.click(); return; }
                        }
                    }
                }""")
            fc = await fc_info.value
            await fc.set_files(upload_path)
            await asyncio.sleep(3)
            log("  ✅ 파일 로드\n")
        except Exception as e:
            log(f"  ❌ 실패: {e}")
            return

        # ① 행1 클릭 [Issue #1 수정 코드]
        log("  [Issue #1] 행1 클릭...")
        clicked_row = await page.evaluate("""() => {
            const tables = document.querySelectorAll('table');
            for (const table of tables) {
                if (table.offsetParent === null) continue;
                const trs = table.querySelectorAll('tr');
                if (trs.length > 2) {
                    const th = trs[1].querySelector('th');
                    if (th && th.textContent.trim() === '1') { th.click(); return true; }
                }
            }
            return false;
        }""")
        log(f"  → {'✅' if clicked_row else '❌'}\n")
        await asyncio.sleep(1)

        # ② 엑셀제목설정
        log("  ② 엑셀제목설정 클릭...")
        await page.evaluate("""() => {
            const btns = document.querySelectorAll('button.WSC_LUXButton');
            for (const btn of btns) {
                if (btn.textContent.trim() === '② 엑셀제목설정') { btn.click(); return; }
            }
        }""")
        await asyncio.sleep(2)

        # ② 엑셀제목설정 확인 모달
        log("  ② 제목설정 확인...")
        await click_modal_button(page, "엑셀제목", "확인")
        await asyncio.sleep(2)

        # 확인 버튼 [Issue #2 수정 코드]
        log("  [Issue #2] 확인 버튼 (수정 코드)...")
        await page.evaluate("""() => {
            const selectors = ['._isDialog', '.LUX_basic_dialog'];
            for (const sel of selectors) {
                for (const dialog of document.querySelectorAll(sel)) {
                    if (dialog.style.display === 'none' || dialog.offsetParent === null) continue;
                    const btns = dialog.querySelectorAll('button.WSC_LUXButton');
                    for (const btn of btns) {
                        if (btn.textContent.trim() === '확인') { btn.click(); return; }
                    }
                }
            }
        }""")
        await asyncio.sleep(3)
        log("  완료\n")

        # 후속 모달 1: 데이터 저장 (#confirm)
        log("  후속 1 → #confirm 확인...")
        await page.evaluate("""() => { const btn = document.querySelector('#confirm'); if (btn) btn.click(); }""")
        await asyncio.sleep(3)

        # 후속 모달 2: 연결되지 않은 사원 안내
        log("  후속 2 → '연결되지 않은 사원' 확인...")
        await click_modal_button(page, "연결되지 않은 사원", "확인")
        await asyncio.sleep(3)

        # 후속 모달 3: 삭제후 업로드 확인 → 취소 (dry_run)
        log("  후속 3 → '삭제후 업로드' 취소...")
        await click_modal_button(page, "삭제후 업로드", "취소")
        await asyncio.sleep(3)

        # 후속 모달 4: 변환 취소 안내
        log("  후속 4 → '변환이 취소' 확인...")
        await click_modal_button(page, "변환이 취소", "확인")
        await asyncio.sleep(2)

        # 에러 감지
        has_error = await page.evaluate("""() => {
            const selectors = ['._isDialog', '.LUX_basic_dialog'];
            for (const sel of selectors) {
                for (const d of document.querySelectorAll(sel)) {
                    if (d.style.display !== 'none' && d.offsetParent !== null) {
                        const text = d.textContent.trim();
                        if (text.includes('오류') || text.includes('실패') || text.includes('에러'))
                            return text.substring(0, 200);
                    }
                }
            }
            return null;
        }""")

        if has_error:
            log(f"\n❌ 에러 감지: {has_error}")
        else:
            log("\n" + "=" * 60)
            log("✅ WEHAGO 전체 자동화 완료! (dry_run: 취소)")
            log("=" * 60)


asyncio.run(main())
