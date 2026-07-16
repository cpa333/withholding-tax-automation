"""SWSA0101 엑셀 처리 모듈

엑셀 다운로드, 업로드 양식 변환, 엑셀 업로드, 모달 핸들러(사원코드연결/제개산).
"""

import asyncio
import os
import sys

from src.automation.wehago._common import (
    log, dismiss_dialogs, click_menu,
    open_collect_menu, close_collect_menu, click_menu_item, _click_modal_text,
)
from src.utils.human import net_mult

if sys.platform == "win32":
    import openpyxl


async def recalculate_salary(page, *, category="고용보험 재계산"):
    """급여자료입력 화면에서 사원 전체 재계산 (엑셀 다운로드 직전 실행).

    WEHAGO SmartA SWSA0101 의 '재계산' 기능: 사원 전체 선택 → 재계산 버튼 →
    지정 항목(기본 '고용보험 재계산') 체크 → 확인(2회: 항목→전체사원) → 결과 모달 닫기.
    라이브 검증(2026-07-16, 리드플렉스).

    ★ 해상도 무관 (다른 PC 모니터와 무관):
      - 전체 선택은 RealGrid JS API (Grids.getActiveGrid().checkAll(true)) 사용 — 좌표 없음.
        Left_grid 포커스는 Playwright locator('#Left_grid canvas') 클릭(요소 live geometry).
      - LUX 버튼/체크박스/확인은 real mouse click(getBoundingClientRect 중심) — JS .click() 무반응.

    Args:
        category: 재계산 항목 텍스트. 해당 행의 체크를 on.
    Returns:
        bool: 재계산 시퀀스 완료 여부. 결과가 "재계산할 사원 없음"이어도 흐름 완료시 True.
    """
    import json as _json

    async def _rclick(js):
        box = await page.evaluate(js)
        if box and box.get("x") is not None:
            await page.mouse.click(box["x"], box["y"])
            return True
        return False

    # 1) 사원 전체 선택 (해상도 무관)
    log("[재계산] 사원 전체 선택...")
    await page.evaluate(
        r"() => { const g = window.Grids && window.Grids.getActiveGrid && window.Grids.getActiveGrid();"
        r" if (g && g.checkAll) g.checkAll(false); }"
    )
    try:
        await page.locator("#Left_grid canvas").first.click(timeout=net_mult(5000))
    except Exception as e:
        log(f"  ⚠ Left_grid 포커스 클릭 실패(무시): {e}")
    await asyncio.sleep(0.5)
    sel = await page.evaluate(
        r"() => { const g = window.Grids && window.Grids.getActiveGrid && window.Grids.getActiveGrid();"
        r" if (!g || !g.checkAll) return null; g.checkAll(true);"
        r" return { total: g.getItemCount ? g.getItemCount() : -1,"
        r"          checked: g.getCheckedRows ? g.getCheckedRows().length : -1 }; }"
    )
    log(f"  전체 선택: {sel}")

    # 2) 재계산 버튼 (#saosnb 내 WSC_LUXButton "재계산")
    log("[재계산] 재계산 버튼 클릭...")
    ok = await _rclick(
        r"() => { const root = document.getElementById('saosnb'); if (!root) return null;"
        r" const btn = Array.from(root.querySelectorAll('button')).find(x => {"
        r" const r = x.getBoundingClientRect(); return r.width > 0 && /재계산/.test(x.textContent || ''); });"
        r" if (!btn) return null; const r = btn.getBoundingClientRect();"
        r" return { x: r.x + r.width/2, y: r.y + r.height/2 }; }"
    )
    if not ok:
        log("  ⚠ 재계산 버튼 없음 — 재계산 스킵")
        return False
    await asyncio.sleep(net_mult(3))

    # 3) 재계산 항목 체크 (해당 행 real click)
    cat_literal = _json.dumps(category)
    log(f"[재계산] 항목 체크: {category}")
    await _rclick(
        r"() => { const label = " + cat_literal + r"; const root = document.getElementById('saosnb') || document;"
        r" const cands = Array.from(root.querySelectorAll('tr, label, .LUXcomp_checkbox, li, td'));"
        r" const target = cands.find(e => { const r = e.getBoundingClientRect();"
        r" return r.width > 0 && r.height > 0 && (e.textContent || '').includes(label); });"
        r" if (!target) return null;"
        r" const cb = target.querySelector('.LUXcomp_checkbox, label, input[type=checkbox], [class*=heckbox]') || target;"
        r" const r = cb.getBoundingClientRect(); return { x: r.x + r.width/2, y: r.y + r.height/2 }; }"
    )
    await asyncio.sleep(net_mult(1))

    # 4) 1차 확인 (보이는 확인 버튼)
    log("[재계산] 1차 확인...")
    await _rclick(
        r"() => { const root = document.getElementById('saosnb');"
        r" const b = Array.from(root.querySelectorAll('button, input[type=button]')).find(el => {"
        r" const r = el.getBoundingClientRect(); return r.width > 0 && (el.textContent || '').trim() === '확인'; });"
        r" if (!b) return null; const r = b.getBoundingClientRect();"
        r" return { x: r.x + r.width/2, y: r.y + r.height/2 }; }"
    )
    await asyncio.sleep(net_mult(3))

    # 5) 2차(최종) 확인 — "전체사원의 자료를 재계산" 다이얼로그
    log("[재계산] 최종 확인(전체사원 재계산)...")
    await _rclick(
        r"() => { const dlgs = Array.from(document.querySelectorAll('._isDialog, .WSC_LUXDraggableDialog'));"
        r" const d = dlgs.find(x => /전체사원의 자료를 재계산/.test(x.textContent || ''));"
        r" if (!d) return null; const b = Array.from(d.querySelectorAll('button, input[type=button]')).find(el => {"
        r" const r = el.getBoundingClientRect(); return r.width > 0 && (el.textContent || '').trim() === '확인'; });"
        r" if (!b) return null; const r = b.getBoundingClientRect();"
        r" return { x: r.x + r.width/2, y: r.y + r.height/2 }; }"
    )

    # 6) 결과 모달 — #confirm 이 나타나면 닫기 (최대 15s 폴링)
    log("[재계산] 결과 모달 대기/처리...")
    dismissed = False
    for _ in range(15):
        cf = await page.evaluate(
            r"() => { const b = document.getElementById('confirm'); if (!b) return null;"
            r" const r = b.getBoundingClientRect(); if (r.width === 0) return null;"
            r" return { x: r.x + r.width/2, y: r.y + r.height/2 }; }"
        )
        if cf:
            await page.mouse.click(cf["x"], cf["y"])
            dismissed = True
            break
        await asyncio.sleep(1)
    await dismiss_dialogs(page)  # 잔여 모달 정리
    log(f"  결과 모달 dismissed={dismissed}")
    return True


async def download_excel(page, save_dir="."):
    """급여자료입력 화면에서 엑셀 다운로드"""
    await close_collect_menu(page)
    log("[엑셀 다운로드] 드롭다운 열기...")
    await open_collect_menu(page)

    download_future = asyncio.Future()

    def on_download(d):
        if not download_future.done():
            log(f"  다운로드 감지: {d.suggested_filename}")
            download_future.set_result(d)

    page.on("download", on_download)

    log("[엑셀 다운로드] 엑셀 내려받기 클릭...")
    await click_menu_item(page, "엑셀 내려받기")

    download = await asyncio.wait_for(download_future, timeout=net_mult(15))
    fname = download.suggested_filename
    save_path = os.path.join(save_dir, fname)
    await download.save_as(save_path)
    log(f"  저장 완료: {save_path}")

    await close_collect_menu(page)
    return os.path.abspath(save_path)


def convert_for_upload(download_path, *, nhis_data=None, nps_member_data=None,
                       nps_retro_data=None, nps_govt_data=None,
                       ei_support_data=None, ei_collect_data=None):
    """다운로드 엑셀을 WEHAGO 업로드 양식으로 변환

    2행 헤더 평탄화, 합계 행 제거.
    사원코드는 WEHAGO 원본 값을 그대로 보존 (zfill 금지).
    텍스트 컬럼(사원코드 등) 셀 서식을 '@'(텍스트)로 통일하여
    WEHAGO 다운로드 엑셀의 서식과 일치시킴.
    raw data(NHIS/NPS)가 제공되면 공제항목에 덮어쓰기.
    """
    wb_src = openpyxl.load_workbook(download_path)
    ws_src = wb_src["Sheet1"]

    headers = []
    for c in range(1, ws_src.max_column + 1):
        h2 = ws_src.cell(2, c).value
        h1 = ws_src.cell(1, c).value
        if h2 and str(h2).strip():
            headers.append(str(h2).strip())
        elif h1 and str(h1).strip():
            headers.append(str(h1).strip())
        else:
            headers.append(None)

    # WEHAGO 다운로드 엑셀에서 텍스트 서식('@')인 컬럼들
    TEXT_COLS = {"사원코드", "사원명", "부서", "직급", "직종"}

    wb_new = openpyxl.Workbook()
    ws_new = wb_new.active
    ws_new.title = "Sheet1"

    for i, header in enumerate(headers, 1):
        ws_new.cell(1, i).value = header
        # 텍스트 컬럼 헤더도 서식 '텍스트'로 통일
        if header in TEXT_COLS:
            ws_new.cell(1, i).number_format = "@"

    new_row = 2
    for r in range(3, ws_src.max_row + 1):
        first_val = ws_src.cell(r, 1).value
        if not first_val or first_val == "합계":
            continue

        for c in range(1, ws_src.max_column + 1):
            val = ws_src.cell(r, c).value
            header = headers[c - 1]

            if header == "사원코드" and val is not None:
                # 사원코드는 WEHAGO 원본 값을 그대로 보존.
                # WEHAGO는 4자리('0001'), 6자리('000008'),
                # 10자리('2019093001'), 1자리('1') 등 다양한 형식을
                # 사용하므로 zfill 등 임의 변환을 금지함.
                val = str(val)

            if val is None:
                val = "" if header in TEXT_COLS else 0

            cell = ws_new.cell(new_row, c)
            cell.value = val
            # 텍스트 컬럼(사원코드 등)은 셀 서식을 '텍스트'로 통일
            if header in TEXT_COLS:
                cell.number_format = "@"
        new_row += 1

    base, ext = os.path.splitext(download_path)
    upload_path = f"{base}_업로드{ext}"
    wb_new.save(upload_path)
    log(f"  변환 완료: {upload_path}")

    # ── Raw data 병합 (옵셔널) ──────────────────────────────────
    # 6개 raw data 중 한 건이라도 있으면 병합 시도 (apply_raw_data 내부 게이트와 동일).
    # 외부 게이트는 apply_raw_data import/호출 오버헤드를 피하기 위한 early-exit.
    # ei_collect(환수금-only)·nps_retro(소급분-only)·nps_govt(국고지원-only) 가
    # 단독으로 주어지는 엣지 케이스도 놓치지 않도록 6인자 모두 검사.
    if (nhis_data or nps_member_data or nps_retro_data
            or nps_govt_data or ei_support_data or ei_collect_data):
        try:
            from src.utils.data_merger import apply_raw_data
            merge_result = apply_raw_data(
                upload_path,
                nhis_data=nhis_data,
                nps_member_data=nps_member_data,
                nps_retro_data=nps_retro_data,
                nps_govt_data=nps_govt_data,
                ei_support_data=ei_support_data,
                ei_collect_data=ei_collect_data,
            )
            log(f"  [원천데이터 반영] NHIS {merge_result.nhis_applied}명, "
                f"NPS {merge_result.nps_applied}명, 고용보험 {merge_result.ei_applied}명"
                f" ({merge_result.employees_matched}명 매칭)")
            for w in merge_result.warnings:
                log(f"  WARN: {w}")
        except Exception as e:
            log(f"  WARN: 원천데이터 병합 실패 (무시): {e}")

    return os.path.abspath(upload_path)


async def _handle_code_link_modal(page):
    """사원코드연결 모달 처리: 변환 → '제외하고 변환됩니다' 확인

    엑셀 파일의 사원과 급여관리 사원이 다를 때 등장.
    파일 선택 직후 또는 후속 모달 처리 중간에 나타날 수 있음.
    """
    for _ in range(3):
        found = await page.evaluate("""() => {
            const all = document.querySelectorAll('*');
            for (const el of all) {
                try {
                    const cs = window.getComputedStyle(el);
                    if ((cs.position !== 'fixed' && cs.position !== 'absolute')
                        || cs.display === 'none' || el.offsetWidth < 50) continue;
                    const z = parseInt(cs.zIndex);
                    if (z < 1000) continue;
                    const txt = el.textContent;
                    if (!txt.includes('사원코드') || !txt.includes('연결')) continue;
                    if (txt.includes('급여대장')) continue;
                    const btns = el.querySelectorAll('button');
                    for (const btn of btns) {
                        if (btn.textContent.trim() === '변환' && btn.offsetWidth > 0) {
                            btn.click(); return 'clicked';
                        }
                    }
                } catch(e) {}
            }
            return null;
        }""")
        if found:
            log("  사원코드연결 → 변환 클릭")
            await asyncio.sleep(2)
            # "연결되지 않은 사원 및 연말 입력된 사원은 제외하고 변환됩니다" → 확인
            await _click_modal_text(page, "제외하고 변환", "확인")
            await asyncio.sleep(1)
        else:
            break


async def _handle_jegasan_modal(page):
    """제개산 모달 처리: 취소 버튼 클릭

    엑셀 업로드 후 데이터 처리 과정에서 특정 수임처에만 등장.
    모달이 없으면 즉시 종료 (불필요한 대기 없음).
    """
    for _ in range(3):
        found = await page.evaluate("""() => {
            const all = document.querySelectorAll('*');
            for (const el of all) {
                try {
                    const cs = window.getComputedStyle(el);
                    if ((cs.position !== 'fixed' && cs.position !== 'absolute')
                        || cs.display === 'none' || el.offsetWidth < 50) continue;
                    const z = parseInt(cs.zIndex);
                    if (z < 1000) continue;
                    if (!el.textContent.includes('제개산')) continue;
                    const btns = el.querySelectorAll('button');
                    for (const btn of btns) {
                        if (btn.textContent.trim() === '취소' && btn.offsetWidth > 0) {
                            btn.click();
                            return 'clicked';
                        }
                    }
                } catch(e) {}
            }
            return null;
        }""")
        if found:
            log("  제개산 모달 → 취소 클릭")
            await asyncio.sleep(1)
        else:
            break


async def _retry_click_dialog_button(page, button_text, max_wait=15):
    """click_dialog_button에 재시도 로직 추가

    느린 환경에서 모달이 늦게 나타나는 경우를 대비해
    max_wait초 동안 0.5초 간격으로 재시도.
    _isDialog / LUX_basic_dialog 셀렉터 + high z-index overlay 모두 탐색.
    """
    attempts = int(max_wait / 0.5)
    for i in range(attempts):
        result = await page.evaluate("""(btnText) => {
            // 1) _isDialog / LUX_basic_dialog
            const selectors = ['._isDialog', '.LUX_basic_dialog'];
            for (const sel of selectors) {
                for (const d of document.querySelectorAll(sel)) {
                    if (d.style.display === 'none' || d.offsetParent === null) continue;
                    const btns = d.querySelectorAll('button, a');
                    for (const b of btns) {
                        if (b.textContent.trim().includes(btnText) && b.offsetWidth > 0) {
                            b.click(); return 'dialog';
                        }
                    }
                }
            }
            // 2) high z-index overlay (폴백)
            const all = document.querySelectorAll('*');
            for (const el of all) {
                try {
                    const cs = window.getComputedStyle(el);
                    if (cs.display === 'none' || el.offsetWidth < 50) continue;
                    const z = parseInt(cs.zIndex);
                    if (z < 1000) continue;
                    if (cs.position !== 'fixed' && cs.position !== 'absolute') continue;
                    const btns = el.querySelectorAll('button');
                    for (const b of btns) {
                        if (b.textContent.trim() === btnText && b.offsetWidth > 0) {
                            b.click(); return 'overlay';
                        }
                    }
                } catch(e) {}
            }
            return null;
        }""", button_text)
        if result:
            log(f"  모달 버튼 클릭: {button_text} ({result}, {(i+1)*0.5:.1f}초 대기)")
            return True
        await asyncio.sleep(0.5)
    return False


async def upload_excel(page, file_path, dry_run=True):
    """변환된 엑셀 파일을 WEHAGO에 업로드"""
    log("[엑셀 업로드] 화면 정리...")
    await dismiss_dialogs(page)

    # 드롭다운 열기
    log("[엑셀 업로드] 드롭다운 열기...")
    await close_collect_menu(page)
    await open_collect_menu(page)

    # 마감 상태 확인: '해제' 버튼이 있으면 마감 완료 → 엑셀 업로드 불가
    status_btn = await page.evaluate("""() => {
        const buttons = document.querySelectorAll(
            '.WSC_LUXTooltip button.WSC_LUXButton, button.WSC_LUXButton'
        );
        for (const btn of buttons) {
            const text = btn.textContent.trim();
            if (['마감', '마감해제', '해제', '완료'].includes(text) && btn.offsetWidth > 0) {
                return text;
            }
        }
        return null;
    }""")
    if status_btn == '해제':
        log("  마감 완료 상태 ('해제' 버튼). 엑셀 업로드를 건너뜁니다.")
        log("[SWSA0101] 업로드 생략 완료")
        return True

    # --- 엑셀 불러오기: 3단계 fallback ---
    log("[엑셀 업로드] 엑셀 불러오기 클릭...")
    file_set = False

    # 1순위: page.mouse.click — 실제 CDP 마우스 이벤트 (신뢰된 사용자 제스처)
    item_rect = await page.evaluate("""() => {
        const menu = document.querySelector('.sao_head_menu');
        if (!menu) return null;
        const items = menu.querySelectorAll('li');
        for (const li of items) {
            if (li.textContent.includes('엑셀 불러오기') && li.offsetHeight > 0) {
                const rect = li.getBoundingClientRect();
                return {
                    x: rect.x + rect.width / 2,
                    y: rect.y + rect.height / 2
                };
            }
        }
        return null;
    }""")

    if item_rect:
        log(f"  항목 위치: ({round(item_rect['x'])}, {round(item_rect['y'])})")
        try:
            async with page.expect_file_chooser(timeout=int(net_mult(15000))) as fc_info:
                await page.mouse.click(item_rect['x'], item_rect['y'])
            file_chooser = await fc_info.value
            log(f"  파일 선택: {file_path}")
            await file_chooser.set_files(file_path)
            file_set = True
        except Exception as e:
            log(f"  mouse.click 파일 선택창 실패: {e}")

    # 2순위: JS evaluate click (기존 방식)
    if not file_set:
        log("[엑셀 업로드] JS evaluate 클릭으로 재시도...")
        try:
            async with page.expect_file_chooser(timeout=int(net_mult(15000))) as fc_info:
                await click_menu_item(page, "엑셀 불러오기")
            file_chooser = await fc_info.value
            log(f"  파일 선택: {file_path}")
            await file_chooser.set_files(file_path)
            file_set = True
        except Exception as e:
            log(f"  JS evaluate 파일 선택창 실패: {e}")

    # 3순위: hidden file input 직접 설정
    if not file_set:
        log("[엑셀 업로드] hidden file input 직접 설정...")
        await click_menu_item(page, "엑셀 불러오기")
        await asyncio.sleep(net_mult(2.0))
        fi_count = await page.evaluate(
            "() => document.querySelectorAll('input[type=\"file\"]').length"
        )
        log(f"  file input 수: {fi_count}")
        if fi_count > 0:
            fi = page.locator('input[type="file"]').first
            await fi.set_input_files(file_path)
            log(f"  파일 설정 완료: {file_path}")
            file_set = True
        else:
            log("  ERROR: file input을 찾지 못해 업로드 불가")
            return False

    await asyncio.sleep(net_mult(3.0))

    # 사원코드연결 모달 (파일 선택 직후 등장 가능)
    await _handle_code_link_modal(page)

    # 제개산 모달 (특정 수임처에서 파일 선택 직후 등장 가능)
    await _handle_jegasan_modal(page)

    # ① 헤더 행(행1) 선택
    log("[엑셀 업로드] ① 헤더 행 선택...")
    clicked = await page.evaluate("""() => {
        const tables = document.querySelectorAll('table');
        for (const table of tables) {
            if (table.offsetParent === null) continue;
            const trs = table.querySelectorAll('tr');
            if (trs.length > 2) {
                const th = trs[1].querySelector('th');
                if (th && th.textContent.trim() === '1') {
                    th.click();
                    return true;
                }
            }
        }
        return false;
    }""")
    if clicked:
        log("  행1 클릭 완료")
    else:
        log("  행1 요소를 찾지 못함")
    await asyncio.sleep(1)

    # ② 엑셀제목설정
    log("[엑셀 업로드] ② 엑셀제목설정 열기...")
    await page.evaluate("""() => {
        const btns = document.querySelectorAll('button.WSC_LUXButton');
        for (const btn of btns) {
            if (btn.textContent.trim() === '② 엑셀제목설정') {
                btn.click();
                return;
            }
        }
    }""")
    await asyncio.sleep(net_mult(2.0))

    log("[엑셀 업로드] ② 제목설정 확인...")
    await _click_modal_text(page, "엑셀제목", "확인")
    await asyncio.sleep(net_mult(2.0))

    # 확인 버튼
    log("[엑셀 업로드] 확인 버튼 클릭...")
    await page.evaluate("""() => {
        const selectors = ['._isDialog', '.LUX_basic_dialog'];
        for (const sel of selectors) {
            for (const dialog of document.querySelectorAll(sel)) {
                if (dialog.style.display === 'none' || dialog.offsetParent === null) continue;
                const btns = dialog.querySelectorAll('button.WSC_LUXButton');
                for (const btn of btns) {
                    if (btn.textContent.trim() === '확인') {
                        btn.click();
                        return;
                    }
                }
            }
        }
    }""")
    await asyncio.sleep(net_mult(5.0))

    # 후속 모달 1: 데이터 저장
    log("[엑셀 업로드] 후속 1/5 → #confirm 확인...")
    await page.evaluate("""() => {
        const btn = document.querySelector('#confirm');
        if (btn) btn.click();
    }""")
    await asyncio.sleep(net_mult(3.0))

    # 제개산 모달 (데이터 저장 후 등장 가능)
    await _handle_jegasan_modal(page)

    # 후속 모달 2: 연결되지 않은 사원
    log("[엑셀 업로드] 후속 2/5 → '연결되지 않은 사원' 확인...")
    await _click_modal_text(page, "연결되지 않은 사원", "확인")
    await asyncio.sleep(net_mult(3.0))

    # 제개산 모달 (사원 연결 처리 후 등장 가능)
    await _handle_jegasan_modal(page)

    # 사원코드연결 모달 (후속 모달 처리 중간에 등장 가능)
    await _handle_code_link_modal(page)

    # 제개산 모달 (사원코드연결 후 등장 가능)
    await _handle_jegasan_modal(page)

    # 후속 모달 3: 삭제후 업로드
    action = "취소" if dry_run else "확인"
    log(f"[엑셀 업로드] 후속 3/5 → '삭제후 업로드' {action}...")
    await _click_modal_text(page, "삭제후 업로드", action)
    await asyncio.sleep(net_mult(3.0))

    # 후속 모달 4: 변환 취소/완료
    if dry_run:
        log("[엑셀 업로드] 후속 4/5 → '변환이 취소' 확인...")
        await _click_modal_text(page, "변환이 취소", "확인")
    else:
        log("[엑셀 업로드] 후속 4/5 → 변환 완료 대기...")
        # 실제 업로드 처리 후 완료 모달이 늦게 나타날 수 있음.
        # 최대 15초 재시도로 "확인" 버튼 탐색 (모달이 없으면 자동 스킵).
        clicked = await _retry_click_dialog_button(page, "확인", max_wait=15)
        if clicked:
            log("[엑셀 업로드] 후속 4/5 → 완료 모달 확인 처리")
        else:
            log("[엑셀 업로드] 후속 4/5 → 완료 모달 없음 (자동 처리됨)")
    await asyncio.sleep(net_mult(2.0))

    # 에러 감지
    has_error = await page.evaluate("""() => {
        const selectors = ['._isDialog', '.LUX_basic_dialog'];
        for (const sel of selectors) {
            for (const d of document.querySelectorAll(sel)) {
                if (d.style.display !== 'none' && d.offsetParent !== null) {
                    const text = d.textContent.trim();
                    if (text.includes('오류') || text.includes('실패') || text.includes('에러')) {
                        return text.substring(0, 300);
                    }
                }
            }
        }
        return null;
    }""")

    if has_error:
        log(f"  업로드 에러 감지: {has_error}")
        return False

    log("  업로드 완료")
    return True
