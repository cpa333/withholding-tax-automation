"""급여자료입력 (SWSA0101) 자동화

엑셀 다운로드 → 업로드 양식 변환 → 엑셀 업로드 → PDF 발급.

사전 조건:
- page가 이미 SmartA 급여 페이지에 있어야 함
- Chrome CDP 모드(port 9223) 실행 상태
"""
import asyncio
import os
import sys
import time

from src.automation.wehago._common import (
    log, dismiss_dialogs, click_menu, click_dialog_button,
    open_collect_menu, close_collect_menu, click_menu_item, _click_modal_text,
    select_dropdown, goto_menu_page, dismiss_print_modals,
)

if sys.platform == "win32":
    import openpyxl
    from pywinauto import Desktop as WinDesktop
    import pywinauto.actionlogger
    pywinauto.actionlogger.ActionLogger.logger.handlers = []

PRINT_DIALOG_TITLE_RE = r"Duzon.*PrintDialog"
PRINT_DIALOG_CLASS_RE = r"WindowsForms10\.Window.*"
SAVE_DIALOG_CLASS = "#32770"
DEFAULT_PRINT_FORMAT = "급여명세(사원당 한장)"


# ═══════════════════════════════════════════════════════════════════════
# SWSA0101 귀속연월 설정용 JS 상수 (React LS_calendar)
# ═══════════════════════════════════════════════════════════════════════

_READ_SWSA_YM_JS = """() => {
    const items = document.querySelectorAll('#SearchMain .item');
    for (const item of items) {
        const title = item.querySelector('.item_title, strong');
        if (title && title.textContent.trim() === '귀속연월') {
            return item.querySelector('.fakeinput')?.textContent.trim() || '';
        }
    }
    return '';
}"""

_READ_CALENDAR_YEAR_JS = """() => {
    return document.querySelector('.LS_calendar .date_day_title')?.textContent.trim() || '';
}"""

_REACT_SET_CALENDAR_YEAR_JS = """(targetYear) => {
    const all = document.querySelectorAll('*');
    for (const el of all) {
        const keys = Object.keys(el).filter(k => k.startsWith('__reactInternalInstance'));
        for (const key of keys) {
            let node = el[key];
            const queue = [node];
            const visited = new Set();
            for (let depth = 0; depth < 25 && queue.length > 0; depth++) {
                const current = queue.shift();
                if (!current || visited.has(current)) continue;
                visited.add(current);
                const inst = current._instance;
                if (inst && inst.state && inst.state.selectedDate
                    && typeof inst.state.selectedDate.year === 'number') {
                    const oldYear = inst.state.selectedDate.year;
                    const oldMonth = inst.state.selectedDate.month;
                    const newMax = {year: targetYear, month: 12};
                    const newMin = inst.state.minDate
                        ? {year: Math.min(inst.state.minDate.year, targetYear - 1), month: 1}
                        : {year: targetYear - 1, month: 1};
                    inst.setState({
                        selectedDate: {year: targetYear, month: oldMonth},
                        maxDate: newMax,
                        minDate: newMin,
                    });
                    return {success: true, oldYear, oldMonth, newMax, newMin};
                }
                if (current._renderedChildren) {
                    for (const child of Object.values(current._renderedChildren)) {
                        if (child) queue.push(child);
                    }
                }
                if (current._renderedComponent) queue.push(current._renderedComponent);
                if (current.child) queue.push(current.child);
                if (current.sibling) queue.push(current.sibling);
                if (current.return) queue.push(current.return);
            }
        }
    }
    return {success: false};
}"""


# ═══════════════════════════════════════════════════════════════════════
# 엑셀 다운로드 / 변환 / 업로드
# ═══════════════════════════════════════════════════════════════════════

async def download_excel(page, save_dir="."):
    """급여자료입력 화면에서 엑셀 다운로드"""
    await close_collect_menu(page)
    log("[엑셀 다운로드] 드롭다운 열기...")
    await open_collect_menu(page)

    download_future = asyncio.Future()

    def on_download(d):
        if not download_future.done():
            log(f"  다운로드 감지: {d.suggested_filename}")
            download_future.set_result(d)

    page.on("download", on_download)

    log("[엑셀 다운로드] 엑셀 내려받기 클릭...")
    await click_menu_item(page, "엑셀 내려받기")

    download = await asyncio.wait_for(download_future, timeout=15)
    fname = download.suggested_filename
    save_path = os.path.join(save_dir, fname)
    await download.save_as(save_path)
    log(f"  저장 완료: {save_path}")

    await close_collect_menu(page)
    return os.path.abspath(save_path)


def convert_for_upload(download_path, *, nhis_data=None, nps_member_data=None,
                       nps_retro_data=None, nps_govt_data=None):
    """다운로드 엑셀을 WEHAGO 업로드 양식으로 변환

    2행 헤더 평탄화, 합계 행 제거, 사원코드 4자리 0-pad.
    텍스트 컬럼(사원코드 등) 셀 서식을 '@'(텍스트)로 통일하여
    WEHAGO 다운로드 엑셀의 서식과 일치시킴.
    raw data(NHIS/NPS)가 제공되면 공제항목에 덮어쓰기.
    """
    wb_src = openpyxl.load_workbook(download_path)
    ws_src = wb_src["Sheet1"]

    headers = []
    for c in range(1, ws_src.max_column + 1):
        h2 = ws_src.cell(2, c).value
        h1 = ws_src.cell(1, c).value
        if h2 and str(h2).strip():
            headers.append(str(h2).strip())
        elif h1 and str(h1).strip():
            headers.append(str(h1).strip())
        else:
            headers.append(None)

    # WEHAGO 다운로드 엑셀에서 텍스트 서식('@')인 컬럼들
    TEXT_COLS = {"사원코드", "사원명", "부서", "직급", "직종"}

    wb_new = openpyxl.Workbook()
    ws_new = wb_new.active
    ws_new.title = "Sheet1"

    for i, header in enumerate(headers, 1):
        ws_new.cell(1, i).value = header
        # 텍스트 컬럼 헤더도 서식 '텍스트'로 통일
        if header in TEXT_COLS:
            ws_new.cell(1, i).number_format = "@"

    new_row = 2
    for r in range(3, ws_src.max_row + 1):
        first_val = ws_src.cell(r, 1).value
        if not first_val or first_val == "합계":
            continue

        for c in range(1, ws_src.max_column + 1):
            val = ws_src.cell(r, c).value
            header = headers[c - 1]

            if header == "사원코드" and isinstance(val, str):
                try:
                    val = str(int(val)).zfill(4)
                except (ValueError, TypeError):
                    pass

            if val is None:
                val = "" if header in TEXT_COLS else 0

            cell = ws_new.cell(new_row, c)
            cell.value = val
            # 텍스트 컬럼(사원코드 등)은 셀 서식을 '텍스트'로 통일
            if header in TEXT_COLS:
                cell.number_format = "@"
        new_row += 1

    base, ext = os.path.splitext(download_path)
    upload_path = f"{base}_업로드{ext}"
    wb_new.save(upload_path)
    log(f"  변환 완료: {upload_path}")

    # ── Raw data 병합 (옵셔널) ──────────────────────────────────
    if nhis_data or nps_member_data:
        try:
            from src.utils.data_merger import apply_raw_data
            merge_result = apply_raw_data(
                upload_path, nhis_data, nps_member_data,
                nps_retro_data, nps_govt_data,
            )
            log(f"  [원천데이터 반영] NHIS {merge_result.nhis_applied}명, NPS {merge_result.nps_applied}명"
                f" ({merge_result.employees_matched}명 매칭)")
            for w in merge_result.warnings:
                log(f"  WARN: {w}")
        except Exception as e:
            log(f"  WARN: 원천데이터 병합 실패 (무시): {e}")

    return os.path.abspath(upload_path)


async def _handle_code_link_modal(page):
    """사원코드연결 모달 처리: 변환 → '제외하고 변환됩니다' 확인

    엑셀 파일의 사원과 급여관리 사원이 다를 때 등장.
    파일 선택 직후 또는 후속 모달 처리 중간에 나타날 수 있음.
    """
    for _ in range(3):
        found = await page.evaluate("""() => {
            const all = document.querySelectorAll('*');
            for (const el of all) {
                try {
                    const cs = window.getComputedStyle(el);
                    if ((cs.position !== 'fixed' && cs.position !== 'absolute')
                        || cs.display === 'none' || el.offsetWidth < 50) continue;
                    const z = parseInt(cs.zIndex);
                    if (z < 1000) continue;
                    const txt = el.textContent;
                    if (!txt.includes('사원코드') || !txt.includes('연결')) continue;
                    if (txt.includes('급여대장')) continue;
                    const btns = el.querySelectorAll('button');
                    for (const btn of btns) {
                        if (btn.textContent.trim() === '변환' && btn.offsetWidth > 0) {
                            btn.click(); return 'clicked';
                        }
                    }
                } catch(e) {}
            }
            return null;
        }""")
        if found:
            log("  사원코드연결 → 변환 클릭")
            await asyncio.sleep(2)
            # "연결되지 않은 사원 및 연말 입력된 사원은 제외하고 변환됩니다" → 확인
            await _click_modal_text(page, "제외하고 변환", "확인")
            await asyncio.sleep(1)
        else:
            break


async def _handle_jegasan_modal(page):
    """제개산 모달 처리: 취소 버튼 클릭

    엑셀 업로드 후 데이터 처리 과정에서 특정 수임처에만 등장.
    모달이 없으면 즉시 종료 (불필요한 대기 없음).
    """
    for _ in range(3):
        found = await page.evaluate("""() => {
            const all = document.querySelectorAll('*');
            for (const el of all) {
                try {
                    const cs = window.getComputedStyle(el);
                    if ((cs.position !== 'fixed' && cs.position !== 'absolute')
                        || cs.display === 'none' || el.offsetWidth < 50) continue;
                    const z = parseInt(cs.zIndex);
                    if (z < 1000) continue;
                    if (!el.textContent.includes('제개산')) continue;
                    const btns = el.querySelectorAll('button');
                    for (const btn of btns) {
                        if (btn.textContent.trim() === '취소' && btn.offsetWidth > 0) {
                            btn.click();
                            return 'clicked';
                        }
                    }
                } catch(e) {}
            }
            return null;
        }""")
        if found:
            log("  제개산 모달 → 취소 클릭")
            await asyncio.sleep(1)
        else:
            break


async def upload_excel(page, file_path, dry_run=True):
    """변환된 엑셀 파일을 WEHAGO에 업로드"""
    log("[엑셀 업로드] 화면 정리...")
    await dismiss_dialogs(page)

    # 드롭다운 열기
    log("[엑셀 업로드] 드롭다운 열기...")
    await close_collect_menu(page)
    await open_collect_menu(page)

    # 마감 상태 확인: '해제' 버튼이 있으면 마감 완료 → 엑셀 업로드 불가
    status_btn = await page.evaluate("""() => {
        const buttons = document.querySelectorAll(
            '.WSC_LUXTooltip button.WSC_LUXButton, button.WSC_LUXButton'
        );
        for (const btn of buttons) {
            const text = btn.textContent.trim();
            if (['마감', '마감해제', '해제', '완료'].includes(text) && btn.offsetWidth > 0) {
                return text;
            }
        }
        return null;
    }""")
    if status_btn == '해제':
        log("  마감 완료 상태 ('해제' 버튼). 엑셀 업로드를 건너뜁니다.")
        log("[SWSA0101] 업로드 생략 완료")
        return True

    # --- 엑셀 불러오기: 3단계 fallback ---
    log("[엑셀 업로드] 엑셀 불러오기 클릭...")
    file_set = False

    # 1순위: page.mouse.click — 실제 CDP 마우스 이벤트 (신뢰된 사용자 제스처)
    item_rect = await page.evaluate("""() => {
        const menu = document.querySelector('.sao_head_menu');
        if (!menu) return null;
        const items = menu.querySelectorAll('li');
        for (const li of items) {
            if (li.textContent.includes('엑셀 불러오기') && li.offsetHeight > 0) {
                const rect = li.getBoundingClientRect();
                return {
                    x: rect.x + rect.width / 2,
                    y: rect.y + rect.height / 2
                };
            }
        }
        return null;
    }""")

    if item_rect:
        log(f"  항목 위치: ({round(item_rect['x'])}, {round(item_rect['y'])})")
        try:
            async with page.expect_file_chooser(timeout=15000) as fc_info:
                await page.mouse.click(item_rect['x'], item_rect['y'])
            file_chooser = await fc_info.value
            log(f"  파일 선택: {file_path}")
            await file_chooser.set_files(file_path)
            file_set = True
        except Exception as e:
            log(f"  mouse.click 파일 선택창 실패: {e}")

    # 2순위: JS evaluate click (기존 방식)
    if not file_set:
        log("[엑셀 업로드] JS evaluate 클릭으로 재시도...")
        try:
            async with page.expect_file_chooser(timeout=15000) as fc_info:
                await click_menu_item(page, "엑셀 불러오기")
            file_chooser = await fc_info.value
            log(f"  파일 선택: {file_path}")
            await file_chooser.set_files(file_path)
            file_set = True
        except Exception as e:
            log(f"  JS evaluate 파일 선택창 실패: {e}")

    # 3순위: hidden file input 직접 설정
    if not file_set:
        log("[엑셀 업로드] hidden file input 직접 설정...")
        await click_menu_item(page, "엑셀 불러오기")
        await asyncio.sleep(2)
        fi_count = await page.evaluate(
            "() => document.querySelectorAll('input[type=\"file\"]').length"
        )
        log(f"  file input 수: {fi_count}")
        if fi_count > 0:
            fi = page.locator('input[type="file"]').first
            await fi.set_input_files(file_path)
            log(f"  파일 설정 완료: {file_path}")
            file_set = True
        else:
            log("  ERROR: file input을 찾지 못해 업로드 불가")
            return False

    await asyncio.sleep(3)

    # 사원코드연결 모달 (파일 선택 직후 등장 가능)
    await _handle_code_link_modal(page)

    # 제개산 모달 (특정 수임처에서 파일 선택 직후 등장 가능)
    await _handle_jegasan_modal(page)

    # ① 헤더 행(행1) 선택
    log("[엑셀 업로드] ① 헤더 행 선택...")
    clicked = await page.evaluate("""() => {
        const tables = document.querySelectorAll('table');
        for (const table of tables) {
            if (table.offsetParent === null) continue;
            const trs = table.querySelectorAll('tr');
            if (trs.length > 2) {
                const th = trs[1].querySelector('th');
                if (th && th.textContent.trim() === '1') {
                    th.click();
                    return true;
                }
            }
        }
        return false;
    }""")
    if clicked:
        log("  행1 클릭 완료")
    else:
        log("  행1 요소를 찾지 못함")
    await asyncio.sleep(1)

    # ② 엑셀제목설정
    log("[엑셀 업로드] ② 엑셀제목설정 열기...")
    await page.evaluate("""() => {
        const btns = document.querySelectorAll('button.WSC_LUXButton');
        for (const btn of btns) {
            if (btn.textContent.trim() === '② 엑셀제목설정') {
                btn.click();
                return;
            }
        }
    }""")
    await asyncio.sleep(2)

    log("[엑셀 업로드] ② 제목설정 확인...")
    await _click_modal_text(page, "엑셀제목", "확인")
    await asyncio.sleep(2)

    # 확인 버튼
    log("[엑셀 업로드] 확인 버튼 클릭...")
    await page.evaluate("""() => {
        const selectors = ['._isDialog', '.LUX_basic_dialog'];
        for (const sel of selectors) {
            for (const dialog of document.querySelectorAll(sel)) {
                if (dialog.style.display === 'none' || dialog.offsetParent === null) continue;
                const btns = dialog.querySelectorAll('button.WSC_LUXButton');
                for (const btn of btns) {
                    if (btn.textContent.trim() === '확인') {
                        btn.click();
                        return;
                    }
                }
            }
        }
    }""")
    await asyncio.sleep(5)

    # 후속 모달 1: 데이터 저장
    log("[엑셀 업로드] 후속 1/5 → #confirm 확인...")
    await page.evaluate("""() => {
        const btn = document.querySelector('#confirm');
        if (btn) btn.click();
    }""")
    await asyncio.sleep(3)

    # 제개산 모달 (데이터 저장 후 등장 가능)
    await _handle_jegasan_modal(page)

    # 후속 모달 2: 연결되지 않은 사원
    log("[엑셀 업로드] 후속 2/5 → '연결되지 않은 사원' 확인...")
    await _click_modal_text(page, "연결되지 않은 사원", "확인")
    await asyncio.sleep(3)

    # 제개산 모달 (사원 연결 처리 후 등장 가능)
    await _handle_jegasan_modal(page)

    # 사원코드연결 모달 (후속 모달 처리 중간에 등장 가능)
    await _handle_code_link_modal(page)

    # 제개산 모달 (사원코드연결 후 등장 가능)
    await _handle_jegasan_modal(page)

    # 후속 모달 3: 삭제후 업로드
    action = "취소" if dry_run else "확인"
    log(f"[엑셀 업로드] 후속 3/5 → '삭제후 업로드' {action}...")
    await _click_modal_text(page, "삭제후 업로드", action)
    await asyncio.sleep(3)

    # 후속 모달 4: 변환 취소/완료
    if dry_run:
        log("[엑셀 업로드] 후속 4/5 → '변환이 취소' 확인...")
        await _click_modal_text(page, "변환이 취소", "확인")
    else:
        log("[엑셀 업로드] 후속 4/5 → 완료 확인...")
        await click_dialog_button(page, "확인")
    await asyncio.sleep(2)

    # 에러 감지
    has_error = await page.evaluate("""() => {
        const selectors = ['._isDialog', '.LUX_basic_dialog'];
        for (const sel of selectors) {
            for (const d of document.querySelectorAll(sel)) {
                if (d.style.display !== 'none' && d.offsetParent !== null) {
                    const text = d.textContent.trim();
                    if (text.includes('오류') || text.includes('실패') || text.includes('에러')) {
                        return text.substring(0, 300);
                    }
                }
            }
        }
        return null;
    }""")

    if has_error:
        log(f"  업로드 에러 감지: {has_error}")
        return False

    log("  업로드 완료")
    return True


# ═══════════════════════════════════════════════════════════════════════
# SWSA0101 귀속연월 설정 (React LS_calendar)
# ═══════════════════════════════════════════════════════════════════════

async def set_swsa_ym(page, year: int, month: int) -> bool:
    """SWSA0101 귀속연월 설정 (React LS_calendar component)

    SWSA0101은 SWTA/SWER와 다른 React 기반 달력(LS_calendar)을 사용.
    Playwright locator.click으로 캘린더 열고, React setState로 연도 변경 후 월 선택.

    Args:
        page: SWSA0101 페이지에 위치한 Playwright page
        year: 목표 연도 (예: 2026)
        month: 목표 월 (1-12)

    Returns:
        True if 귀속연월 설정 성공, False otherwise
    """
    from src.automation.wehago._common import _safe_evaluate

    target_ym = f"{year}.{month:02d}"

    for attempt in range(3):
        log(f"    [귀속연월] 시도 {attempt+1}/3: {target_ym}")

        # ── 현재 값 읽기 ──────────────────────────────────────
        cur_ym = await _safe_evaluate(page, _READ_SWSA_YM_JS)
        if cur_ym == target_ym:
            log(f"    [귀속연월] 이미 {target_ym} — 스킵")
            return True

        log(f"    [귀속연월] 현재: {cur_ym} → 목표: {target_ym}")

        # ── 캘린더 열기 (반드시 Playwright click — JS evaluate는 합성 이벤트) ──
        try:
            await page.locator(
                "#SearchMain .item:first-child .fakebutton"
            ).click(timeout=5000)
            await asyncio.sleep(1)
        except Exception as e:
            log(f"    [귀속연월] 캘린더 열기 실패: {e}")
            await asyncio.sleep(1)
            continue

        # ── 연도 확인 및 React setState ──────────────────────
        cal_yr_text = await _safe_evaluate(page, _READ_CALENDAR_YEAR_JS)
        if not cal_yr_text:
            log("    [귀속연월] 캘린더 연도 읽기 실패")
            await asyncio.sleep(1)
            continue

        try:
            cal_yr = int(cal_yr_text)
        except (ValueError, TypeError):
            cal_yr = None

        if cal_yr is not None and cal_yr != year:
            log(f"    [귀속연월] React setState: {cal_yr} → {year}")
            result = await _safe_evaluate(
                page, _REACT_SET_CALENDAR_YEAR_JS, year,
            )
            if not result or not result.get("success"):
                log(f"    [귀속연월] React setState 실패: {result}")
                await asyncio.sleep(1)
                continue
            await asyncio.sleep(1)

            # 연도 변경 확인
            new_cal_yr = await _safe_evaluate(page, _READ_CALENDAR_YEAR_JS)
            if new_cal_yr != str(year):
                log(f"    [귀속연월] 연도 변경 확인 실패: {new_cal_yr}")
                await asyncio.sleep(1)
                continue

        # ── 월 클릭 ──────────────────────────────────────────
        try:
            month_btn = page.locator(
                f'.LS_calendar td.date_day button:has-text("{month}월")'
            )
            await month_btn.first.click(timeout=3000)
            await asyncio.sleep(1)
        except Exception as e:
            log(f"    [귀속연월] {month}월 클릭 실패: {e}")
            await asyncio.sleep(1)
            continue

        # ── 최종 검증 ────────────────────────────────────────
        final_ym = await _safe_evaluate(page, _READ_SWSA_YM_JS)
        if final_ym == target_ym:
            log(f"    [귀속연월] 설정 완료: {target_ym}")
            return True

        log(f"    [귀속연월] 검증 실패: {final_ym} (예상: {target_ym})")
        await asyncio.sleep(1)

    log(f"    [귀속연월] 3회 재시도 후 실패")
    return False


# ═══════════════════════════════════════════════════════════════════════
# PDF 다운로드 (OS-level PrintDialog 제어)
# ═══════════════════════════════════════════════════════════════════════

async def open_print_dialog(page):
    """브라우저에서 #print 버튼 → 일괄출력 메뉴 클릭하여 PrintDialog 실행"""
    log("[PDF] #print 버튼 클릭...")
    await page.evaluate("""() => {
        const btn = document.querySelector('#print');
        if (btn) btn.click();
    }""")
    await asyncio.sleep(1)

    log("[PDF] 일괄출력 메뉴 클릭...")
    await click_menu_item(page, "일괄출력")

    if sys.platform != "win32":
        log("  Windows 전용 기능입니다.")
        return False

    log("[PDF] PrintDialog 대기...")
    for i in range(15):
        await asyncio.sleep(2)
        if _print_dialog_exists():
            log("  PrintDialog 열림 확인")
            return True
        if i % 3 == 2:
            log(f"  대기 중... {(i+1)*2}초")

    log("  PrintDialog 열림 타임아웃")
    return False


def _find_print_dialog():
    desktop = WinDesktop(backend='uia')
    return desktop.window(title_re=PRINT_DIALOG_TITLE_RE, class_name_re=PRINT_DIALOG_CLASS_RE)


def _print_dialog_exists():
    try:
        dlg = _find_print_dialog()
        return dlg.exists(timeout=1)
    except Exception:
        return False


def _close_existing_print_dialog():
    """기존 PrintDialog가 떠 있으면 종료 + 경고 모달 처리"""
    if not _print_dialog_exists():
        return

    log("  기존 PrintDialog 감지. 정리 중...")

    try:
        dlg = _find_print_dialog()
        for btn in dlg.descendants(control_type='Button'):
            name = btn.element_info.element.CurrentName
            if name and name == '확인':
                btn.click_input()
                log("  경고 모달 '확인' 클릭")
                time.sleep(1)
                break
    except Exception:
        pass

    try:
        time.sleep(1)
        dlg = _find_print_dialog()
        dlg.child_window(auto_id='btnClose', control_type='Button').click_input()
        log("  기존 PrintDialog 종료")
        time.sleep(2)
    except Exception:
        pass


def _select_print_format(target_text):
    """PrintDialog 인쇄형태 드롭다운에서 항목 선택"""
    dlg = _find_print_dialog()
    dlg.set_focus()
    time.sleep(0.5)

    cb = dlg.child_window(auto_id='cbContents', control_type='ComboBox')
    open_btn = cb.children(control_type='Button')[0]
    open_btn.click_input()
    time.sleep(1.5)

    items = cb.descendants(control_type='ListItem')
    for item in items:
        name = item.element_info.element.CurrentName
        if name and target_text in name:
            item.click_input()
            log(f"  인쇄형태 선택: {name}")
            time.sleep(2)
            return True

    log(f"  인쇄형태 '{target_text}' 항목을 찾지 못함")
    return False


def _click_save_pdf():
    """PrintDialog에서 PDF 저장 버튼 클릭"""
    dlg = _find_print_dialog()
    btn = dlg.child_window(auto_id='btnSavePDF', control_type='Button')
    btn.click_input()
    log("  PDF 저장 버튼 클릭")
    time.sleep(3)


def _handle_save_dialog(save_path):
    """Windows '다른 이름으로 저장' 대화상자에서 경로 입력 후 저장"""
    desktop = WinDesktop(backend='win32')
    dlg = desktop.window(title='다른 이름으로 저장', class_name=SAVE_DIALOG_CLASS)

    edit = dlg.child_window(class_name='Edit')
    edit.set_edit_text(save_path)
    time.sleep(1)

    save_btn = dlg.child_window(title='저장(&S)', class_name='Button')
    save_btn.click_input()
    time.sleep(3)

    if os.path.exists(save_path) and os.path.getsize(save_path) > 0:
        log(f"  PDF 저장 완료: {save_path} ({os.path.getsize(save_path):,} bytes)")
        return True

    log("  PDF 파일 저장 실패")
    return False


def _close_print_dialog():
    """PrintDialog 종료"""
    dlg = _find_print_dialog()
    dlg.child_window(auto_id='btnClose', control_type='Button').click_input()
    log("  PrintDialog 종료")


async def download_pdf(page, save_dir, print_format=DEFAULT_PRINT_FORMAT):
    """PrintDialog를 통해 PDF 다운로드"""
    if sys.platform != "win32":
        log("  PDF 다운로드는 Windows 전용 기능입니다.")
        return None

    loop = asyncio.get_event_loop()

    # 기존 PrintDialog 정리
    await loop.run_in_executor(None, _close_existing_print_dialog)

    # PrintDialog 실행
    if not await open_print_dialog(page):
        return None

    # 인쇄형태 선택
    selected = await loop.run_in_executor(None, _select_print_format, print_format)
    if not selected:
        return None

    # PDF 버튼
    await loop.run_in_executor(None, _click_save_pdf)

    # Windows 저장 대화상자
    os.makedirs(save_dir, exist_ok=True)
    filename = f"{time.strftime('%Y%m%d_%H%M%S')}_{print_format.split('(')[0]}.pdf"
    save_path = os.path.join(save_dir, filename)

    saved = await loop.run_in_executor(None, _handle_save_dialog, save_path)
    if not saved:
        return None

    # PrintDialog 종료
    await loop.run_in_executor(None, _close_print_dialog)

    return os.path.abspath(save_path)


# ═══════════════════════════════════════════════════════════════════════
# 메인 플로우
# ═══════════════════════════════════════════════════════════════════════

async def run_swsa0101(page, save_dir, *, dry_run=True, year=None, month=None):
    """급여자료입력 전체 자동화

    Args:
        page: SmartA 급여 페이지에 위치한 Playwright page
        save_dir: 파일 저장 디렉토리
        dry_run: True면 업로드 후 취소(테스트), False면 확인(실운영)
        year: 귀속연도 (None이면 이전 달 자동 계산)
        month: 귀속월 (None이면 이전 달 자동 계산)
    """
    from src.automation.wehago._common import (
        navigate_to_swsa0101, compute_target_period,
    )

    # year/month 기본값: 이전 달
    if year is None or month is None:
        year, month = compute_target_period()

    # [1] SWSA0101 메뉴 이동 + 귀속연월 설정
    log(f"[SWSA0101] 급여자료입력 메뉴 이동 ({year}.{month:02d})...")
    ok = await navigate_to_swsa0101(page, year=year, month=month)
    if not ok:
        log("[SWSA0101] 이동/설정 실패")
        return

    # [4] 엑셀 다운로드
    log("[SWSA0101] 엑셀 다운로드...")
    download_path = await download_excel(page, save_dir)

    # [5] 업로드 양식 변환
    log("[SWSA0101] 업로드 양식 변환...")
    upload_path = convert_for_upload(download_path)

    # [5.5] 업로드 전 모달 정리
    log("[SWSA0101] 업로드 전 화면 정리...")
    await dismiss_dialogs(page)

    # [6] 엑셀 업로드
    log("[SWSA0101] 엑셀 업로드...")
    success = await upload_excel(page, upload_path, dry_run=dry_run)
    if success:
        log("  업로드 완료!")
    else:
        log("  업로드 중 에러 발생. 화면을 확인하세요.")

    # [7] PDF 다운로드
    log("[SWSA0101] PDF 다운로드...")
    pdf_path = await download_pdf(page, save_dir)
    if pdf_path:
        log(f"  PDF 완료: {pdf_path}")
    else:
        log("  PDF 다운로드 실패")

    # 급여대장 일괄인쇄 모달 닫기
    log("[SWSA0101] 일괄인쇄 모달 정리...")
    await dismiss_print_modals(page)

    log("[SWSA0101] 완료")


# ═══════════════════════════════════════════════════════════════════════
# 독립 실행
# ═══════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    import io
    if sys.platform == "win32":
        sys.stdout = io.TextIOWrapper(sys.stdout.detach(), encoding='utf-8')
        sys.stderr = io.TextIOWrapper(sys.stderr.detach(), encoding='utf-8')

    async def _main():
        from playwright.async_api import async_playwright
        from src.utils.chrome_cdp import launch_chrome, connect_page
        from src.automation.wehago._common import wait_for_login, goto_salary_page

        company = input("수임처 이름: ").strip()
        if not company:
            print("수임처 이름이 필요합니다.")
            return

        year_input = input("연도 (Enter=자동): ").strip()
        month_input = input("월 (Enter=자동): ").strip()
        year = int(year_input) if year_input else None
        month = int(month_input) if month_input else None

        launch_chrome()
        async with async_playwright() as p:
            browser, context, page = await connect_page(p)
            if not await wait_for_login(page):
                return
            await dismiss_dialogs(page)
            if not await goto_salary_page(page, company):
                return
            await dismiss_dialogs(page)

            save_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", "results"))
            await run_swsa0101(page, save_dir, dry_run=True, year=year, month=month)

    asyncio.run(_main())
