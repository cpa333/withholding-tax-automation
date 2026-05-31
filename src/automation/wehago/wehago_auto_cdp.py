"""CDP로 기존 Chrome에 연결하여 WEHAGO 급여 자동화"""
import asyncio
import sys
import os
import subprocess
import time
from playwright.async_api import async_playwright
import openpyxl
if sys.platform == "win32":
    from pywinauto import Desktop as WinDesktop
    import pywinauto.actionlogger
    pywinauto.actionlogger.ActionLogger.logger.handlers = []

if sys.platform == "win32":
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.detach(), encoding='utf-8')
    sys.stderr = io.TextIOWrapper(sys.stderr.detach(), encoding='utf-8')

# PROJECT_ROOT to sys.path for src.* imports
PROJECT_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
sys.path.insert(0, PROJECT_ROOT)

from src.utils.chrome_cdp import CDP_URL, CDP_PORT
from src.utils.log import log

WEHAGO_URL = "https://www.wehago.com/"
SMARTA_BASE = "https://smarta.wehago.com"


def find_chrome():
    """Chrome 실행 파일 경로 찾기"""
    paths = [
        os.path.join(os.environ.get("PROGRAMFILES", ""), "Google", "Chrome", "Application", "chrome.exe"),
        os.path.join(os.environ.get("PROGRAMFILES(X86)", ""), "Google", "Chrome", "Application", "chrome.exe"),
        os.path.join(os.environ.get("LOCALAPPDATA", ""), "Google", "Chrome", "Application", "chrome.exe"),
    ]
    for p in paths:
        if os.path.exists(p):
            return p
    return None


async def check_cdp_available():
    """CDP 포트가 활성인지 확인"""
    try:
        import urllib.request
        with urllib.request.urlopen(f"{CDP_URL}/json/version", timeout=2) as resp:
            return resp.status == 200
    except Exception:
        return False


def kill_chrome():
    """기존 Chrome 프로세스 종료"""
    subprocess.run(["taskkill", "/F", "/IM", "chrome.exe"],
                    stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
    time.sleep(3)


async def launch_chrome():
    """Chrome을 디버깅 모드로 실행 (사용자 프로필 사용)"""
    if await check_cdp_available():
        log("CDP 포트 이미 활성. 기존 Chrome에 연결합니다.")
        return True

    chrome_path = find_chrome()
    if not chrome_path:
        log("Chrome 브라우저를 찾을 수 없습니다.")
        return False

    log("기존 Chrome을 종료하고 디버깅 모드로 재실행합니다...")
    kill_chrome()

    user_data = os.path.join(os.environ.get("LOCALAPPDATA", ""),
                             "Google", "Chrome", "User Data")

    subprocess.Popen([
        chrome_path,
        f"--remote-debugging-port={CDP_PORT}",
        "--user-data-dir=" + user_data,
        "--profile-directory=Profile 2",
        "--start-maximized",
        WEHAGO_URL,
    ], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)

    log("Chrome 실행 대기...")
    for i in range(30):
        await asyncio.sleep(1)
        if await check_cdp_available():
            log("Chrome 연결 성공!")
            return True

    log("Chrome 실행 실패.")
    return False


async def connect_browser(playwright):
    """CDP로 Chrome에 연결하고 첫 페이지 반환"""
    from src.utils.stealth import stealth_all_pages, register_auto_stealth

    browser = await playwright.chromium.connect_over_cdp(CDP_URL)
    context = browser.contexts[0]

    await stealth_all_pages(context)
    register_auto_stealth(context)

    page = context.pages[0] if context.pages else await context.new_page()
    return browser, context, page


async def wait_for_login(page):
    """WEHAGO 로그인 완료 대기 (수동 로그인)"""
    await page.goto(WEHAGO_URL + "#/main", wait_until="domcontentloaded")
    await asyncio.sleep(2)

    # 로그인 상태 확인: 메인 페이지에 수임처 리스트가 보이면 로그인됨
    if await page.locator("#company_").count() > 0 or await page.locator("text=나의 수임처").count() > 0:
        log("이미 로그인되어 있습니다.")
        return True

    log("\n브라우저에서 WEHAGO 로그인을 진행해 주세요.")
    log("로그인 완료 후 Enter를 누르세요.")

    # 로그인 대기: 수임처 리스트가 나타날 때까지 폴링
    for _ in range(120):
        await asyncio.sleep(5)
        try:
            if await page.locator("text=나의 수임처").count() > 0:
                log("로그인 확인됨.")
                return True
            await page.reload(wait_until="domcontentloaded")
            await asyncio.sleep(2)
            if await page.locator("text=나의 수임처").count() > 0:
                log("로그인 확인됨.")
                return True
        except Exception:
            pass

    log("로그인 대기 시간 초과 (10분).")
    return False


async def get_client_salary_url(page, company_name):
    """수임처의 급여(SmartA) URL 획득

    WEHAGO 메인의 수임처 카드에서 window.open을 가로채어
    급여 버튼 클릭 시 열리는 SmartA URL을 캡처합니다.
    """
    # window.open 가로채기
    await page.evaluate("""() => {
        window.__capturedUrl = null;
        window.__origOpen = window.open;
        window.open = function(url) {
            window.__capturedUrl = url;
            return null;
        };
    }""")

    # 수임처 카드에서 급여 버튼 클릭
    clicked = await page.evaluate("""(companyName) => {
        const allDivs = document.querySelectorAll('[id^="company_"]');
        for (const div of allDivs) {
            const nameEl = div.querySelector('a');
            if (nameEl && nameEl.textContent.trim() === companyName) {
                let card = div;
                for (let i = 0; i < 3; i++) card = card.parentElement;
                const buttons = card.querySelectorAll('button.btn_quick');
                for (const btn of buttons) {
                    if (btn.querySelector('span')?.textContent.trim() === '급여') {
                        btn.click();
                        return true;
                    }
                }
            }
        }
        return false;
    }""", company_name)

    if not clicked:
        log(f"수임처 '{company_name}'의 급여 버튼을 찾지 못했습니다.")
        return None

    # 캡처된 URL 확인
    await asyncio.sleep(1)
    url = await page.evaluate("() => window.__capturedUrl")

    # window.open 복원
    await page.evaluate("() => { window.open = window.__origOpen; }")

    return url


async def goto_salary_page(page, company_name):
    """수임처의 SmartA 급여 메인 페이지로 이동"""
    url = await get_client_salary_url(page, company_name)
    if not url:
        return False

    log(f"SmartA 급여 URL: {url}")
    await page.goto(url, wait_until="domcontentloaded", timeout=30000)

    # SPA 로드 완료 대기: 사이드 메뉴가 나타날 때까지
    for i in range(15):
        await asyncio.sleep(2)
        if await page.locator("a.text_link").count() > 0:
            break
    else:
        log("  SmartA 페이지 로드 타임아웃")
        return False

    log(f"페이지 이동 완료: {await page.title()}")
    return True


async def click_menu(page, menu_id):
    """SmartA 메인의 사이드 메뉴 클릭 (SPA 내부 라우팅)

    menu_id 예시: 'SWSA0101' (급여자료입력), 'SWSA0104' (은행별이체명세) 등
    """
    await page.evaluate("""(menuId) => {
        const link = document.querySelector('a#' + menuId + '.text_link');
        if (link) link.click();
    }""", menu_id)
    await asyncio.sleep(3)
    log(f"메뉴 이동 완료: {await page.title()}")


async def goto_menu_page(page, menu_id):
    """SmartA 내 다른 메뉴 페이지로 이동 (URL 해시 교체)

    현재 SmartA URL에서 메뉴 ID 해시만 교체하여 페이지 전환.
    사이드 메뉴에 보이지 않는 메뉴(예: SWTA0101 원천징수이행상황신고서)도 이동 가능.
    """
    current_url = page.url
    # 현재 URL에서 /SWSA0101 같은 메뉴 해시를 찾아 교체
    import re
    new_url = re.sub(r'/[A-Z]+\d+(?=[?#]|$)', '/' + menu_id, current_url)
    if new_url == current_url:
        log(f"  URL 교체 실패: {menu_id}")
        return False

    log(f"메뉴 이동: {menu_id}")
    await page.goto(new_url, wait_until="domcontentloaded", timeout=30000)
    await asyncio.sleep(2)
    log(f"이동 완료: {await page.title()}")
    return True


async def get_report_period_type(page):
    """원천징수이행상황신고서의 매월/반기 라디오 상태 반환

    매월/반기 라디오를 찾아 checked 상태를 반환.
    둘 다 체크 안 되어 있으면 기본값으로 매월을 선택 후 반환.
    """
    result = await page.evaluate("""() => {
        const radios = document.querySelectorAll('input.LSinput[type=radio]');
        const monthlyRadios = [];
        for (const r of radios) {
            const label = r.closest('label')?.querySelector('.label_text')?.textContent?.trim();
            if (label === '매월' || label === '반기') {
                monthlyRadios.push({radio: r, label, checked: r.checked});
            }
        }
        // 이미 체크된 것 반환
        const checked = monthlyRadios.find(r => r.checked);
        if (checked) return checked.label;
        // 체크된 게 없으면 매월 선택
        const monthly = monthlyRadios.find(r => r.label === '매월');
        if (monthly) {
            monthly.radio.click();
            return '매월';
        }
        return null;
    }""")
    return result


async def set_period_fields(page, year, start_month, end_month):
    """지급기간/귀속기간 설정 (#SearchMain 내 .item)

    각 기간 항목: div[tabindex=0] × 4 (시작년도, 시작월표시, 종료년도, 종료월표시)
    연도: triple-click → 타이핑 → Enter (WSC_LUXAlert 차단 주의)
    월: 화살표 버튼 클릭 → 드롭다운에서 li 선택
    """
    # WSC_LUXAlert 오버레이가 클릭 차단하므로 먼저 닫기
    await page.evaluate("""() => {
        document.querySelectorAll('.WSC_LUXAlert').forEach(a => {
            const btn = a.querySelector('button.WSC_LUXButton');
            if (btn) btn.click();
            a.style.display = 'none';
        });
    }""")

    period_labels = ['귀속기간', '지급기간']

    # JS에서 직접 rect를 가져와서 좌표 클릭 (JSHandle 불안정 회피)
    # 기간 항목만 필터링: title에 '기간' 포함 + tabindex div 4개 + sprite 버튼 2개
    rects = await page.evaluate("""() => {
        const results = [];
        const items = document.querySelectorAll('#SearchMain .item');
        items.forEach((item, idx) => {
            const title = item.querySelector('.item_title, strong');
            const titleText = title ? title.textContent.trim() : '';
            if (!titleText.includes('기간')) return;
            const inputDivs = item.querySelectorAll('div[tabindex="0"]');
            const spriteBtns = item.querySelectorAll('button .WSC_LUXSpriteIcon');
            if (inputDivs.length < 4 || spriteBtns.length < 2) return;
            const entry = {idx, title: titleText, years: [], months: []};
            inputDivs.forEach((d, i) => {
                const r = d.getBoundingClientRect();
                entry.years.push({i, text: d.textContent.trim(), x: r.x, y: r.y, w: r.width, h: r.height});
            });
            spriteBtns.forEach((s, i) => {
                const btn = s.closest('button');
                const r = btn.getBoundingClientRect();
                entry.months.push({i, x: r.x, y: r.y, w: r.width, h: r.height});
            });
            results.push(entry);
        });
        return results;
    }""")

    for idx, rect in enumerate(rects):
        label = rect['title'] if rect['title'] else (period_labels[idx] if idx < 2 else f'항목{idx}')
        log(f"  {label}: {year}년 {start_month:02d}월 ~ {year}년 {end_month:02d}월")

        # 시작 연도 (triple-click으로 텍스트 전체 선택)
        if len(rect['years']) > 0:
            y = rect['years'][0]
            await page.mouse.click(y['x'] + y['w'] / 2, y['y'] + y['h'] / 2, click_count=3)
            await asyncio.sleep(0.3)
            await page.keyboard.type(str(year))
            await page.keyboard.press('Enter')
            await asyncio.sleep(0.3)

        # 시작 월
        if len(rect['months']) > 0:
            m = rect['months'][0]
            await page.mouse.click(m['x'] + m['w'] / 2, m['y'] + m['h'] / 2)
            await asyncio.sleep(0.5)
            target_text = f"{start_month:02d}"
            clicked = await page.evaluate(f"""() => {{
                const lis = document.querySelectorAll('div[style*="position: fixed"] li div');
                for (const li of lis) {{
                    if (li.textContent.trim() === '{target_text}') {{
                        li.click(); return true;
                    }}
                }}
                return false;
            }}""")
            if not clicked:
                log(f"    시작월 {target_text} 선택 실패")
            await asyncio.sleep(0.3)

        # 종료 연도
        if len(rect['years']) > 2:
            y = rect['years'][2]
            await page.mouse.click(y['x'] + y['w'] / 2, y['y'] + y['h'] / 2, click_count=3)
            await asyncio.sleep(0.3)
            await page.keyboard.type(str(year))
            await page.keyboard.press('Enter')
            await asyncio.sleep(0.3)

        # 종료 월
        if len(rect['months']) > 1:
            m = rect['months'][1]
            await page.mouse.click(m['x'] + m['w'] / 2, m['y'] + m['h'] / 2)
            await asyncio.sleep(0.5)
            target_text = f"{end_month:02d}"
            clicked = await page.evaluate(f"""() => {{
                const lis = document.querySelectorAll('div[style*="position: fixed"] li div');
                for (const li of lis) {{
                    if (li.textContent.trim() === '{target_text}') {{
                        li.click(); return true;
                    }}
                }}
                return false;
            }}""")
            if not clicked:
                log(f"    종료월 {target_text} 선택 실패")
            await asyncio.sleep(0.3)

        # 연도 검증 및 재시도
        verify = await page.evaluate(f"""() => {{
            const items = document.querySelectorAll('#SearchMain .item');
            if (!items[{idx}]) return null;
            const divs = items[{idx}].querySelectorAll('div[tabindex="0"]');
            return Array.from(divs).map(d => d.textContent.trim());
        }}""")
        if verify and verify[0] != str(year):
            log(f"    시작 연도 재시도 ({verify[0]} -> {year})...")
            y = rect['years'][0]
            await page.mouse.click(y['x'] + y['w'] / 2, y['y'] + y['h'] / 2, click_count=3)
            await asyncio.sleep(0.3)
            await page.keyboard.type(str(year))
            await page.keyboard.press('Enter')
            await asyncio.sleep(0.3)

    return True


async def select_dropdown(page, dropdown_index, option_text):
    """커스텀 드롭다운(LS_ngh_select2)에서 옵션 선택

    Args:
        dropdown_index: LS_ngh_select2 요소 중 몇 번째인지 (0-based)
        option_text: 선택할 옵션 텍스트 (부분 매치)
    """
    # 드롭다운 열기
    await page.evaluate("""(idx) => {
        const dd = document.querySelectorAll('.LS_ngh_select2')[idx];
        if (dd) dd.querySelector('.LSbutton').click();
    }""", dropdown_index)
    await asyncio.sleep(1)

    # 옵션 선택
    await page.evaluate("""(args) => {
        const items = document.querySelectorAll('.LSselectResult li');
        for (const li of items) {
            if (li.textContent.includes(args.text)) {
                li.querySelector('a').click();
                return true;
            }
        }
        return false;
    }""", {"text": option_text})
    await asyncio.sleep(1)

    # 선택값 확인
    value = await page.evaluate("""(idx) => {
        const dd = document.querySelectorAll('.LS_ngh_select2')[idx];
        return dd ? dd.querySelector('.fakeinput').textContent.trim() : '';
    }""", dropdown_index)
    log(f"드롭다운 선택: {value}")


async def _click_modal_text(page, text_fragment, action):
    """특정 텍스트가 포함된 모달에서 action(확인/취소) 버튼 클릭"""
    for _ in range(20):
        result = await page.evaluate("""(args) => {
            const all = document.querySelectorAll('*');
            for (const el of all) {
                if (!el.textContent.includes(args.fragment)) continue;
                const btns = el.querySelectorAll('button');
                for (const btn of btns) {
                    if (btn.textContent.trim() === args.action && btn.offsetWidth > 0) {
                        btn.click();
                        return args.action;
                    }
                }
            }
            return null;
        }""", {"fragment": text_fragment, "action": action})
        if result:
            return True
        await asyncio.sleep(0.5)
    return False


async def click_dialog_button(page, button_text):
    """현재 떠 있는 모달/다이얼로그에서 지정된 텍스트의 버튼 클릭"""
    await page.evaluate("""(btnText) => {
        const selectors = ['._isDialog', '.LUX_basic_dialog'];
        let target = null;
        for (const sel of selectors) {
            for (const d of document.querySelectorAll(sel)) {
                if (d.style.display !== 'none') { target = d; break; }
            }
            if (target) break;
        }
        if (!target) return;
        const btns = target.querySelectorAll('button, a');
        for (const b of btns) {
            if (b.textContent.trim().includes(btnText)) { b.click(); return; }
        }
    }""", button_text)
    await asyncio.sleep(1)
    log(f"  모달 버튼 클릭: {button_text}")


async def dismiss_dialogs(page):
    """표시 중인 팝업/다이얼로그가 있으면 모두 닫기

    WEHAGO/SmartA의 다이얼로그를 탐지하여 닫습니다.
    대상: _isDialog, LUX_basic_dialog + z-index >= 1000 fixed 오버레이
    1) '닫기' 텍스트 버튼 → 2) X 버튼 → 3) '확인' 버튼 → 4) '취소' 버튼 순으로 닫습니다.
    팝업이 없으면 아무것도 하지 않으며, 중첩 시 모두 사라질 때까지 반복합니다.
    """
    for _ in range(20):
        closed = await page.evaluate("""() => {
            // 1) 명시적 다이얼로그 탐색 (display/visibility)
            const selectors = ['._isDialog', '.LUX_basic_dialog'];
            let target = null;
            for (const sel of selectors) {
                for (const d of document.querySelectorAll(sel)) {
                    const cs = window.getComputedStyle(d);
                    if (cs.display !== 'none' && cs.visibility !== 'hidden'
                        && d.offsetParent !== null && d.offsetWidth > 0) {
                        target = d; break;
                    }
                }
                if (target) break;
            }

            // 2) z-index >= 1000 fixed 오버레이 (텍스트 있는 것만, Snackbar 제외)
            if (!target) {
                const all = document.querySelectorAll('*');
                for (const el of all) {
                    const cs = window.getComputedStyle(el);
                    if (cs.position !== 'fixed' || cs.display === 'none'
                        || parseInt(cs.zIndex) < 1000 || el.offsetWidth < 100) continue;
                    if (el.classList.contains('WSC_LUXSnackbar')) continue;
                    if (el.textContent.trim().length === 0) continue;
                    target = el; break;
                }
            }

            if (!target) return null;

            const allBtns = target.querySelectorAll('button, a');

            // 1) 닫기
            for (const btn of allBtns) {
                if (btn.textContent.trim() === '닫기') { btn.click(); return '닫기'; }
            }
            // 2) X (빈 텍스트 WSC_LUXButton)
            const luxBtns = target.querySelectorAll('button.WSC_LUXButton');
            for (const btn of luxBtns) {
                if (!btn.textContent.trim()) { btn.click(); return 'X'; }
            }
            // 3) 확인 in dialog_btnbx
            const confirmBtn = target.querySelector('.dialog_btnbx button');
            if (confirmBtn) { confirmBtn.click(); return '확인(btnbx)'; }
            // 4) 확인
            for (const btn of allBtns) {
                if (btn.textContent.trim() === '확인' && btn.offsetWidth > 0) {
                    btn.click(); return '확인';
                }
            }
            // 5) 취소
            for (const btn of allBtns) {
                if (btn.textContent.trim() === '취소') { btn.click(); return '취소'; }
            }
            return 'stuck';
        }""")
        if not closed:
            return
        log(f"  팝업 닫음 ({closed})")
        await asyncio.sleep(0.5)


async def open_collect_menu(page):
    """우측 끝 #collect 버튼 클릭하여 드롭다운 메뉴 열기"""
    await page.evaluate("""() => {
        const btn = document.querySelector('#collect');
        if (btn) btn.click();
    }""")
    await asyncio.sleep(1)


async def click_menu_item(page, item_text):
    """sao_head_menu 드롭다운에서 특정 텍스트 항목의 a 태그 클릭"""
    return await page.evaluate("""(text) => {
        const menu = document.querySelector('.sao_head_menu');
        if (!menu) return false;
        const items = menu.querySelectorAll('li');
        for (const li of items) {
            if (li.textContent.includes(text)) {
                const a = li.querySelector('a');
                if (a) { a.click(); return true; }
                li.click();
                return true;
            }
        }
        return false;
    }""", item_text)


async def download_excel(page, save_dir="."):
    """급여자료입력 화면에서 엑셀 다운로드

    #collect 드롭다운 → 엑셀 내려받기 클릭 후 파일 저장.
    저장된 파일의 절대경로를 반환.
    """
    log("[엑셀 다운로드] 드롭다운 열기...")
    await open_collect_menu(page)

    download_future = asyncio.Future()

    def on_download(d):
        if not download_future.done():
            log(f"  다운로드 감지: {d.suggested_filename}")
            download_future.set_result(d)

    page.on("download", on_download)

    log("[엑셀 다운로드] 엑셀 내려받기 클릭...")
    await click_menu_item(page, "엑셀 내려받기")

    download = await asyncio.wait_for(download_future, timeout=15)
    fname = download.suggested_filename
    save_path = os.path.join(save_dir, fname)
    await download.save_as(save_path)
    log(f"  저장 완료: {save_path}")
    return os.path.abspath(save_path)


def convert_for_upload(download_path):
    """다운로드 엑셀을 WEHAGO 업로드 양식으로 변환

    다운로드 파일의 2행 헤더(행1 대분류 + 행2 세부항목)를
    단일 헤더 행으로 평탄화하여 모든 열을 보존.
    사원코드는 4자리 0-pad 문자열로 변환 (예: "0005").
    """
    wb_src = openpyxl.load_workbook(download_path)
    ws_src = wb_src["Sheet1"]

    # 헤더 평탄화: 행2(세부항목) 값 우선, 없으면 행1(대분류) 값 사용
    headers = []
    for c in range(1, ws_src.max_column + 1):
        h2 = ws_src.cell(2, c).value
        h1 = ws_src.cell(1, c).value
        if h2 and str(h2).strip():
            headers.append(str(h2).strip())
        elif h1 and str(h1).strip():
            headers.append(str(h1).strip())
        else:
            headers.append(None)

    TEXT_COLS = {"사원코드", "사원명", "부서", "직급", "직종"}

    wb_new = openpyxl.Workbook()
    ws_new = wb_new.active
    ws_new.title = "Sheet1"

    # 헤더 행
    for i, header in enumerate(headers, 1):
        ws_new.cell(1, i).value = header

    # 데이터 행 (행3 ~ 마지막, 합계 제외)
    new_row = 2
    for r in range(3, ws_src.max_row + 1):
        first_val = ws_src.cell(r, 1).value
        if not first_val or first_val == "합계":
            continue

        for c in range(1, ws_src.max_column + 1):
            val = ws_src.cell(r, c).value
            header = headers[c - 1]

            if header == "사원코드" and isinstance(val, str):
                try:
                    val = str(int(val)).zfill(4)
                except (ValueError, TypeError):
                    pass

            if val is None:
                val = "" if header in TEXT_COLS else 0

            ws_new.cell(new_row, c).value = val
        new_row += 1

    base, ext = os.path.splitext(download_path)
    upload_path = f"{base}_업로드{ext}"
    wb_new.save(upload_path)
    log(f"  변환 완료: {upload_path}")
    return os.path.abspath(upload_path)


async def upload_excel(page, file_path, dry_run=True):
    """변환된 엑셀 파일을 WEHAGO에 업로드

    #collect 드롭다운 → 엑셀 불러오기 → file chooser →
    ① 헤더 행 선택 → ② 엑셀제목설정 확인 → 확인 →
    후속 모달에서 dry_run=True면 취소, False면 확인.
    """
    log("[엑셀 업로드] 드롭다운 열기...")
    await open_collect_menu(page)

    log("[엑셀 업로드] 엑셀 불러오기 클릭...")
    async with page.expect_file_chooser(timeout=10000) as fc_info:
        await click_menu_item(page, "엑셀 불러오기")

    file_chooser = await fc_info.value
    log(f"  파일 선택: {file_path}")
    await file_chooser.set_files(file_path)
    await asyncio.sleep(3)

    # ① 엑셀내역 테이블에서 헤더 행(행1) 선택
    # JS element.click() 사용 — CDP 좌표는 DPR/배율에 따라 빗나감
    log("[엑셀 업로드] ① 헤더 행 선택...")
    clicked = await page.evaluate("""() => {
        const tables = document.querySelectorAll('table');
        for (const table of tables) {
            if (table.offsetParent === null) continue;
            const trs = table.querySelectorAll('tr');
            if (trs.length > 2) {
                const th = trs[1].querySelector('th');
                if (th && th.textContent.trim() === '1') {
                    th.click();
                    return true;
                }
            }
        }
        return false;
    }""")
    if clicked:
        log("  행1 클릭 완료")
    else:
        log("  행1 요소를 찾지 못함")
    await asyncio.sleep(1)

    # ② 엑셀제목설정 버튼 클릭
    log("[엑셀 업로드] ② 엑셀제목설정 열기...")
    await page.evaluate("""() => {
        const btns = document.querySelectorAll('button.WSC_LUXButton');
        for (const btn of btns) {
            if (btn.textContent.trim() === '② 엑셀제목설정') {
                btn.click();
                return;
            }
        }
    }""")
    await asyncio.sleep(2)

    # ② 엑셀제목설정 확인 모달
    log("[엑셀 업로드] ② 제목설정 확인...")
    await _click_modal_text(page, "엑셀제목", "확인")
    await asyncio.sleep(2)

    # 확인 버튼 클릭 (현재 표시 중인 다이얼로그 내부의 버튼)
    log("[엑셀 업로드] 확인 버튼 클릭...")
    await page.evaluate("""() => {
        const selectors = ['._isDialog', '.LUX_basic_dialog'];
        for (const sel of selectors) {
            for (const dialog of document.querySelectorAll(sel)) {
                if (dialog.style.display === 'none' || dialog.offsetParent === null) continue;
                const btns = dialog.querySelectorAll('button.WSC_LUXButton');
                for (const btn of btns) {
                    if (btn.textContent.trim() === '확인') {
                        btn.click();
                        return;
                    }
                }
            }
        }
    }""")
    await asyncio.sleep(5)

    # 후속 모달 1: 데이터 저장 확인 (#confirm)
    log("[엑셀 업로드] 후속 모달 1/5 → #confirm 확인...")
    await page.evaluate("""() => {
        const btn = document.querySelector('#confirm');
        if (btn) btn.click();
    }""")
    await asyncio.sleep(3)

    # 후속 모달 2: 연결되지 않은 사원 안내
    log("[엑셀 업로드] 후속 모달 2/5 → '연결되지 않은 사원' 확인...")
    await _click_modal_text(page, "연결되지 않은 사원", "확인")
    await asyncio.sleep(3)

    # 후속 모달 3: 삭제후 업로드 안내 (dry_run=True → 취소, False → 확인)
    action = "취소" if dry_run else "확인"
    log(f"[엑셀 업로드] 후속 모달 3/5 → '삭제후 업로드' {action}...")
    await _click_modal_text(page, "삭제후 업로드", action)
    await asyncio.sleep(3)

    # 후속 모달 4: 변환 취소/완료 안내
    if dry_run:
        log("[엑셀 업로드] 후속 모달 4/5 → '변환이 취소' 확인...")
        await _click_modal_text(page, "변환이 취소", "확인")
    else:
        log("[엑셀 업로드] 후속 모달 4/5 → 완료 확인...")
        await click_dialog_button(page, "확인")
    await asyncio.sleep(2)

    # 업로드 후 에러 감지
    has_error = await page.evaluate("""() => {
        const selectors = ['._isDialog', '.LUX_basic_dialog'];
        for (const sel of selectors) {
            for (const d of document.querySelectorAll(sel)) {
                if (d.style.display !== 'none' && d.offsetParent !== null) {
                    const text = d.textContent.trim();
                    if (text.includes('오류') || text.includes('실패') || text.includes('에러')) {
                        return text.substring(0, 300);
                    }
                }
            }
        }
        return null;
    }""")

    if has_error:
        log(f"  업로드 에러 감지: {has_error}")
        return False

    log("  업로드 완료")
    return True


# ===== PDF 다운로드 (OS-level PrintDialog 제어) =====

PRINT_DIALOG_TITLE = "Duzon - PrintDialog"
PRINT_DIALOG_CLASS = "WindowsForms10.Window.8.app.0.141b42a_r8_ad1"
SAVE_DIALOG_CLASS = "#32770"
DEFAULT_PRINT_FORMAT = "급여명세(사원당 한장)"


async def open_print_dialog(page):
    """브라우저에서 #print 버튼 → 일괄출력 메뉴 클릭하여 PrintDialog 실행"""
    log("[PDF] #print 버튼 클릭...")
    await page.evaluate("""() => {
        const btn = document.querySelector('#print');
        if (btn) btn.click();
    }""")
    await asyncio.sleep(1)

    log("[PDF] 일괄출력 메뉴 클릭...")
    await click_menu_item(page, "일괄출력")

    # PrintDialog가 열릴 때까지 대기 (최대 30초)
    if sys.platform != "win32":
        log("  Windows 전용 기능입니다.")
        return False

    log("[PDF] PrintDialog 대기...")
    for i in range(15):
        await asyncio.sleep(2)
        if _print_dialog_exists():
            log("  PrintDialog 열림 확인")
            return True
        if i % 3 == 2:
            log(f"  대기 중... {(i+1)*2}초")

    log("  PrintDialog 열림 타임아웃")
    return False


def _find_print_dialog():
    """OS 레벨에서 PrintDialog 창 찾기 (pywinauto, auto_id 기반)"""
    desktop = WinDesktop(backend='uia')
    return desktop.window(title_re=PRINT_DIALOG_TITLE, class_name=PRINT_DIALOG_CLASS)


def _print_dialog_exists():
    """PrintDialog가 현재 떠 있는지 확인"""
    try:
        dlg = _find_print_dialog()
        return dlg.exists(timeout=1)
    except Exception:
        return False


def _close_existing_print_dialog():
    """기존 PrintDialog가 떠 있으면 종료 + '이미 인쇄함이 있습니다' 모달 처리

    PrintDialog가 열려 있는 상태에서 다시 출력을 시도하면
    '이미 셸함수에 있는 인쇄함이 있습니다' 경고 모달이 나타남.
    이 모달의 '확인' 버튼을 누르고 기존 PrintDialog를 종료.
    """
    if not _print_dialog_exists():
        return

    log("  기존 PrintDialog 감지. 정리 중...")

    # 1) 경고 모달(WindowsForms 모달) 처리: "확인" 버튼 클릭
    try:
        dlg = _find_print_dialog()
        for btn in dlg.descendants(control_type='Button'):
            name = btn.element_info.element.CurrentName
            if name and name == '확인':
                btn.click_input()
                log("  경고 모달 '확인' 클릭")
                time.sleep(1)
                break
    except Exception:
        pass

    # 2) 기존 PrintDialog 종료
    try:
        time.sleep(1)
        dlg = _find_print_dialog()
        dlg.child_window(auto_id='btnClose', control_type='Button').click_input()
        log("  기존 PrintDialog 종료")
        time.sleep(2)
    except Exception:
        pass


def _select_print_format(target_text):
    """PrintDialog의 인쇄형태 드롭다운에서 항목 선택 (auto_id + CurrentName 기반)"""
    dlg = _find_print_dialog()
    dlg.set_focus()
    time.sleep(0.5)

    cb = dlg.child_window(auto_id='cbContents', control_type='ComboBox')

    # 열기 버튼: cb 자식 중 첫 번째 Button
    open_btn = cb.children(control_type='Button')[0]
    open_btn.click_input()
    time.sleep(1.5)

    # ListItem 탐색 후 CurrentName으로 매치
    items = cb.descendants(control_type='ListItem')
    for item in items:
        name = item.element_info.element.CurrentName
        if name and target_text in name:
            item.click_input()
            log(f"  인쇄형태 선택: {name}")
            time.sleep(2)
            return True

    log(f"  인쇄형태 '{target_text}' 항목을 찾지 못함")
    return False


def _click_save_pdf():
    """PrintDialog에서 PDF 저장 버튼 클릭"""
    dlg = _find_print_dialog()
    btn = dlg.child_window(auto_id='btnSavePDF', control_type='Button')
    btn.click_input()
    log("  PDF 저장 버튼 클릭")
    time.sleep(3)


def _handle_save_dialog(save_path):
    """Windows '다른 이름으로 저장' 대화상자에서 경로 입력 후 저장"""
    desktop = WinDesktop(backend='win32')
    dlg = desktop.window(title='다른 이름으로 저장', class_name=SAVE_DIALOG_CLASS)

    # 파일명 Edit에 경로 입력
    edit = dlg.child_window(class_name='Edit')
    edit.set_edit_text(save_path)
    time.sleep(1)

    # 저장 버튼 클릭
    save_btn = dlg.child_window(title='저장(&S)', class_name='Button')
    save_btn.click_input()
    time.sleep(3)

    if os.path.exists(save_path) and os.path.getsize(save_path) > 0:
        log(f"  PDF 저장 완료: {save_path} ({os.path.getsize(save_path):,} bytes)")
        return True

    log("  PDF 파일 저장 실패")
    return False


def _close_print_dialog():
    """PrintDialog 종료 (btnClose 버튼)"""
    dlg = _find_print_dialog()
    dlg.child_window(auto_id='btnClose', control_type='Button').click_input()
    log("  PrintDialog 종료")


async def download_pdf(page, save_dir, print_format=DEFAULT_PRINT_FORMAT):
    """PrintDialog를 통해 PDF 다운로드

    브라우저 #print 버튼 → 일괄출력 → OS PrintDialog →
    인쇄형태 선택 → PDF 저장 → Windows 저장 대화상자 → PrintDialog 종료

    Args:
        page: Playwright page 객체
        save_dir: PDF 저장 디렉토리
        print_format: 인쇄형태 드롭다운에서 선택할 항목 텍스트 (부분 매치)
    """
    if sys.platform != "win32":
        log("  PDF 다운로드는 Windows 전용 기능입니다.")
        return None

    # 0. 기존 PrintDialog 정리 (떠 있으면 경고 모달 처리 후 종료)
    loop = asyncio.get_event_loop()
    await loop.run_in_executor(None, _close_existing_print_dialog)

    # 1. 브라우저에서 PrintDialog 실행
    if not await open_print_dialog(page):
        return None

    # 2. 인쇄형태 선택
    selected = await loop.run_in_executor(None, _select_print_format, print_format)
    if not selected:
        return None

    # 3. PDF 버튼 클릭
    await loop.run_in_executor(None, _click_save_pdf)

    # 4. Windows 저장 대화상자 처리
    os.makedirs(save_dir, exist_ok=True)
    filename = f"{time.strftime('%Y%m%d_%H%M%S')}_{print_format.split('(')[0]}.pdf"
    save_path = os.path.join(save_dir, filename)

    saved = await loop.run_in_executor(None, _handle_save_dialog, save_path)
    if not saved:
        return None

    # 5. PrintDialog 종료
    await loop.run_in_executor(None, _close_print_dialog)

    return os.path.abspath(save_path)


async def run(dry_run=True):
    """전체 자동화 실행

    Args:
        dry_run: True면 업로드 후 취소(개발용), False면 확인(실제 운영용)
    """
    async with async_playwright() as p:
        # ===== [1] Chrome 실행 =====
        log("[1/15] Chrome 실행...")
        if not await launch_chrome():
            return

        # ===== [2] 연결, 로그인, 팝업 닫기 =====
        log("[2/15] Chrome 연결 및 로그인 확인...")
        browser, context, page = await connect_browser(p)
        if not await wait_for_login(page):
            return
        await dismiss_dialogs(page)

        # ===== [3] 수임처 급여 페이지 이동 =====
        log("[3/15] 수임처 급여 페이지 이동...")
        company_name = "[테스트] (주)리틀치프코리아"
        if not await goto_salary_page(page, company_name):
            return
        await dismiss_dialogs(page)

        # ===== [4] 급여자료입력 메뉴 이동 =====
        log("[4/15] 급여자료입력 메뉴 이동...")
        await click_menu(page, "SWSA0101")
        await asyncio.sleep(3)

        # 간이세액 개정 안내 모달 닫기 (X 버튼)
        log("[4-1] 간이세액 안내 모달 닫기...")
        await page.evaluate("""() => {
            const all = document.querySelectorAll('*');
            for (const el of all) {
                const cs = window.getComputedStyle(el);
                if (cs.position !== 'fixed' || cs.display === 'none' ||
                    parseInt(cs.zIndex) <= 100 || el.offsetWidth <= 100) continue;
                if (!el.textContent.includes('간이세액')) continue;
                const btns = el.querySelectorAll('button.WSC_LUXButton');
                for (const btn of btns) {
                    if (!btn.textContent.trim() && btn.offsetWidth > 0) { btn.click(); return; }
                }
            }
        }""")
        await asyncio.sleep(1)
        await dismiss_dialogs(page)

        # ===== [5] 구분 드롭다운: 급여+상여 선택 =====
        log("[5/15] 구분 드롭다운 → 급여+상여 선택...")
        await select_dropdown(page, 0, "급여+상여")

        # ===== [6-7] 복사후 재계산 모달 (조건부) =====
        await asyncio.sleep(1)
        has_modal = await page.evaluate("""() => {
            const selectors = ['._isDialog', '.LUX_basic_dialog'];
            for (const sel of selectors) {
                for (const d of document.querySelectorAll(sel)) {
                    if (d.style.display !== 'none') return true;
                }
            }
            return false;
        }""")
        if has_modal:
            log("[6/14] 복사후 재계산 버튼 클릭...")
            await click_dialog_button(page, "복사후 재계산")
            await asyncio.sleep(1)

            log("[7/14] 확인 모달 → 취소 클릭...")
            await click_dialog_button(page, "취소")
        else:
            log("[6-7/15] 모달 없음 - 스킵")

        # ===== [8] 엑셀 다운로드 =====
        log("[8/15] 엑셀 다운로드...")
        save_dir = os.path.dirname(os.path.abspath(__file__))
        save_dir = os.path.abspath(os.path.join(save_dir, "..", "..", "results"))
        os.makedirs(save_dir, exist_ok=True)
        download_path = await download_excel(page, save_dir)

        # ===== [9] 업로드 양식 변환 =====
        log("[9/15] 업로드 양식 변환...")
        upload_path = convert_for_upload(download_path)

        # ===== [10] 엑셀 업로드 =====
        log("[10/15] 엑셀 업로드...")
        success = await upload_excel(page, upload_path, dry_run=dry_run)

        if success:
            log(f"\n'{company_name}' 급여자료 엑셀 업로드 완료!")
        else:
            log(f"\n'{company_name}' 업로드 중 에러 발생. 화면을 확인하세요.")
        log(f"URL: {page.url}")

        # ===== [11-14] PDF 다운로드 =====
        log("[11/15] #print 버튼 → 일괄출력 클릭...")
        log("[12/15] 인쇄형태 선택...")
        log("[13/15] PDF 저장...")
        log("[14/15] PrintDialog 종료...")
        pdf_path = await download_pdf(page, save_dir)

        if pdf_path:
            log(f"\nPDF 다운로드 완료: {pdf_path}")
        else:
            log("\nPDF 다운로드 실패.")

        # ===== [15] 원천징수이행상황신고서 페이지 이동 =====
        log("[15/15] 원천징수이행상황신고서(SWTA0101) 이동...")
        await goto_menu_page(page, "SWTA0101")
        await asyncio.sleep(3)

        # 모달 확인 후 닫기
        await dismiss_dialogs(page)

        # ===== [15-1] 매월/반기 확인 → 귀속기간/지급기간 설정 =====
        from datetime import datetime
        now = datetime.now()
        period_type = await get_report_period_type(page)
        log(f"  신고유형: {period_type}")

        if period_type == "매월":
            # 저번달
            if now.month == 1:
                target_year = now.year - 1
                target_month = 12
            else:
                target_year = now.year
                target_month = now.month - 1
            log(f"  매월 → {target_year}년 {target_month:02d}월")
            await set_period_fields(page, target_year, target_month, target_month)
        elif period_type == "반기":
            target_year = now.year
            log(f"  반기 → {target_year}년 01월 ~ 06월")
            await set_period_fields(page, target_year, 1, 6)
        else:
            log(f"  알 수 없는 신고유형: {period_type}")

        # ===== [15-2] 조회 버튼 클릭 =====
        log("  조회 버튼 클릭...")
        await page.evaluate("""() => {
            const btns = document.querySelectorAll('#Search button');
            for (const btn of btns) {
                if (btn.textContent.trim() === '조회' && btn.getBoundingClientRect().width > 0) {
                    btn.click();
                    return true;
                }
            }
            return false;
        }""")
        await asyncio.sleep(3)

        # ===== [15-3] 마감/마감해제 버튼 처리 =====
        btn_text = await page.evaluate("""() => {
            const btns = document.querySelectorAll('.WSC_LUXTooltip button.WSC_LUXButton');
            for (const btn of btns) {
                const text = btn.textContent.trim();
                if (text === '마감' || text === '마감해제') return text;
            }
            return null;
        }""")
        if btn_text == "마감":
            log("  마감 버튼 클릭 (마감해제)...")
            await page.evaluate("""() => {
                const btns = document.querySelectorAll('.WSC_LUXTooltip button.WSC_LUXButton');
                for (const btn of btns) {
                    if (btn.textContent.trim() === '마감') { btn.click(); return; }
                }
            }""")
            await asyncio.sleep(1)
        elif btn_text == "마감해제":
            log("  이미 마감해제 상태 - 스킵")
        else:
            log(f"  마감 버튼 상태: {btn_text}")

        # ===== [16] 원천징수 전자신고(SWER0101) 이동 =====
        log("[16] 원천징수 전자신고(SWER0101) 이동...")
        await goto_menu_page(page, "SWER0101")
        await asyncio.sleep(3)

        # 모달 확인 후 닫기 (제출자등록 안내 등)
        log("  모달 확인...")
        await dismiss_dialogs(page)

        # ===== [17] 지급기간 설정 =====
        log("[17] 지급기간 설정...")
        if now.month == 1:
            target_year = now.year - 1
            target_month = 12
        else:
            target_year = now.year
            target_month = now.month - 1
        log(f"  지급기간: {target_year}년 {target_month:02d}월")
        await set_period_fields(page, target_year, target_month, target_month)

        # ===== [18] 수임처 아이콘 클릭 → 확인 =====
        log("[18] 수임처 아이콘 클릭...")
        await page.evaluate("""() => {
            const items = document.querySelectorAll('#SearchMain .item');
            for (const item of items) {
                const title = item.querySelector('.item_title, strong');
                if (!title || !title.textContent.includes('수임처')) continue;
                const btns = item.querySelectorAll('button.WSC_LUXButton');
                for (const btn of btns) {
                    const r = btn.getBoundingClientRect();
                    if (r.width > 0 && r.height > 0 && !btn.textContent.trim()) {
                        btn.click(); return;
                    }
                }
            }
        }""")
        await asyncio.sleep(2)

        # 회사 코드도움 모달에서 확인(enter) 클릭
        log("  회사 코드도움 확인 클릭...")
        for frame in page.frames:
            try:
                await frame.evaluate("""() => {
                    const all = document.querySelectorAll('*');
                    for (const el of all) {
                        try {
                            const cs = window.getComputedStyle(el);
                            const z = parseInt(cs.zIndex) || 0;
                            if (z < 1000 || cs.display === 'none' || el.offsetWidth < 100) continue;
                            if (!el.textContent.includes('코드도움')) continue;
                            const btns = el.querySelectorAll('button');
                            for (const btn of btns) {
                                if (btn.textContent.trim() === '확인(enter)' && btn.offsetWidth > 0) {
                                    btn.click(); return;
                                }
                            }
                        } catch(e) {}
                    }
                }""")
            except Exception:
                pass
        await asyncio.sleep(2)

        # ===== [19] 제작(F4) 버튼 클릭 =====
        log("[19] 제작(F4) 버튼 클릭...")
        await page.evaluate("""() => {
            const all = document.querySelectorAll('button.WSC_LUXButton');
            for (const btn of all) {
                if (btn.textContent.trim() === '제작(F4)') {
                    const r = btn.getBoundingClientRect();
                    if (r.y < 100) { btn.click(); return; }
                }
            }
        }""")
        await asyncio.sleep(2)

        # 제작제외 참고사항 모달 닫기 (iframe 내부 _isDialog에서 확인 클릭)
        log("  제작제외 참고사항 모달 닫기...")
        for frame in page.frames:
            try:
                await frame.evaluate("""() => {
                    const dialogs = document.querySelectorAll('._isDialog');
                    for (const d of dialogs) {
                        if (!d.textContent.includes('참고사항')) continue;
                        const btns = d.querySelectorAll('button');
                        for (const btn of btns) {
                            const txt = btn.textContent.trim();
                            if ((txt === '확인(enter)' || txt === '확인') && btn.offsetWidth > 0) {
                                btn.click(); return;
                            }
                        }
                    }
                }""")
            except Exception:
                pass
        await asyncio.sleep(2)


if __name__ == "__main__":
    asyncio.run(run())
