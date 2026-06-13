"""WEHAGO 엑셀에 raw data 병합 모듈

NHIS/NPS raw data를 사원명 기준으로 매칭하여 WEHAGO 업로드 엑셀의
공제항목(건강보험, 요양보험, 국민연금 등)에 덮어쓰기.
"""
import logging
import re
from dataclasses import dataclass, field

import openpyxl

from src.utils.raw_data_reader import NhisEmployeeData

logger = logging.getLogger(__name__)


# ═══════════════════════════════════════════════════════════════════════
# 컬럼명 유연 매칭
# ═══════════════════════════════════════════════════════════════════════

COLUMN_ALIASES: dict[str, list[str]] = {
    "건강보험": ["건강보험"],
    "장기요양보험료": ["장기요양보험료", "요양보험", "장기요양"],
    # 정산 — 3가지 네이밍 패턴 지원:
    #   Simple:   건강보험료정산, 장기요양보험료정산
    #   Standard: 건강보험정산, 요양보험정산
    #   Detailed: 연말정산(건강보험), 퇴직정산(건강보험) — 별도 컬럼
    "건강보험정산": ["건강보험정산", "건강보험료정산"],
    "장기요양보험정산": ["장기요양보험정산", "요양보험정산", "장기요양보험료정산"],
    "연말정산_건강": ["연말정산(건강보험)", "연말정산(건강)"],
    "연말정산_요양": ["연말정산(장기요양보험)", "연말정산(요양)", "연말정산(장기요양)"],
    "퇴직정산_건강": ["퇴직정산(건강보험)", "퇴직정산(건강)"],
    "퇴직정산_요양": ["퇴직정산(장기요양보험)", "퇴직정산(요양)", "퇴직정산(장기요양)"],
    "국민연금": ["국민연금"],
    "국민연금정산": ["국민연금정산", "국민연금 소급분"],
}


@dataclass
class MergeResult:
    """병합 결과"""
    path: str
    employees_matched: int = 0
    employees_unmatched: list[str] = field(default_factory=list)
    nhis_applied: int = 0
    nps_applied: int = 0
    warnings: list[str] = field(default_factory=list)


def _normalize(s) -> str:
    """매칭용 정규화: 모든 공백 제거.

    WEHAGO 다운로드 템플릿이 '건강보험 정산', '장기요양보험 정산'처럼
    공백을 포함한 컬럼명을 사용하므로, alias(공백 없음)와 비교 전에 통일.
    """
    return re.sub(r"\s+", "", str(s))


def _find_column_index(headers: list[str], target_key: str) -> int | None:
    """헤더 리스트에서 target_key의 alias와 매칭되는 컬럼 인덱스(0-based) 반환.

    헤더·alias 양쪽을 공백 제거한 뒤 ① 정확 일치 우선(오매칭 방지)
    ② 부분문자열(괄호/접미사 변형 대비) 순으로 탐색.
    """
    aliases = COLUMN_ALIASES.get(target_key, [target_key])
    norm_aliases = [_normalize(a) for a in aliases]
    norm_headers = [_normalize(h) for h in headers]

    # 1패스: 정확 일치 (정규화 후) — short alias가 더 긴 컬럼을 오인하는 경합 방지
    for i, nh in enumerate(norm_headers):
        if nh and nh in norm_aliases:
            return i
    # 2패스: 부분문자열 (정규화 후)
    for i, nh in enumerate(norm_headers):
        if not nh:
            continue
        if any(a and a in nh for a in norm_aliases):
            return i
    return None


def _find_name_column(headers: list[str]) -> int | None:
    """사원명 컬럼 인덱스(0-based) 탐색"""
    for i, h in enumerate(headers):
        if h and "사원명" in str(h):
            return i
    return None


def apply_raw_data(
    upload_path: str,
    nhis_data: dict[str, NhisEmployeeData] | None = None,
    nps_member_data: dict[str, int] | None = None,
    nps_retro_data: dict[str, int] | None = None,
    nps_govt_data: dict[str, int] | None = None,
) -> MergeResult:
    """WEHAGO 업로드 엑셀에 raw data 덮어쓰기 (사원명 기준 매칭)

    Args:
        upload_path: convert_for_upload()에서 생성한 엑셀 경로
        nhis_data: {성명: NhisEmployeeData} or None
        nps_member_data: {성명: 근로자기여금} or None
        nps_retro_data: {성명: 본인기여금} or None
        nps_govt_data: {성명: 국고지원금액(전액)/2} or None

    Returns:
        MergeResult
    """
    result = MergeResult(path=upload_path)

    if not nhis_data and not nps_member_data and not nps_retro_data and not nps_govt_data:
        result.warnings.append("반영할 raw data 없음")
        return result

    try:
        wb = openpyxl.load_workbook(upload_path)
    except Exception as e:
        result.warnings.append(f"엑셀 열기 실패: {e}")
        return result

    ws = wb.active

    # 헤더 읽기 (R1)
    headers = []
    for c in range(1, ws.max_column + 1):
        val = ws.cell(1, c).value
        headers.append(str(val).strip() if val else "")

    # 필수 컬럼 탐색
    name_col = _find_name_column(headers)
    if name_col is None:
        result.warnings.append("'사원명' 컬럼을 찾을 수 없음")
        wb.close()
        return result

    # 매핑 대상 컬럼 인덱스 탐색
    col_map: dict[str, int | None] = {}
    for key in COLUMN_ALIASES:
        col_idx = _find_column_index(headers, key)
        col_map[key] = col_idx  # None이면 해당 컬럼이 WEHAGO 엑셀에 없음

    missing_cols = [k for k, v in col_map.items() if v is None]
    if missing_cols:
        result.warnings.append(f"매핑 불가 컬럼: {missing_cols}")

    # 행별 처리 (R2~)
    wehago_names_seen = set()
    for r in range(2, ws.max_row + 1):
        emp_name = ws.cell(r, name_col + 1).value
        if not emp_name or not str(emp_name).strip():
            continue
        emp_name = str(emp_name).strip()
        wehago_names_seen.add(emp_name)

        # --- NHIS 데이터 반영 ---
        nhis_emp = nhis_data.get(emp_name) if nhis_data else None
        if nhis_emp:
            _apply_nhis_row(ws, r, col_map, nhis_emp)
            result.nhis_applied += 1

        # --- NPS 데이터 반영 ---
        nps_amount = nps_member_data.get(emp_name) if nps_member_data else None
        nps_retro = nps_retro_data.get(emp_name) if nps_retro_data else None
        nps_govt = nps_govt_data.get(emp_name) if nps_govt_data else None

        if nps_amount is not None or nps_retro is not None or nps_govt is not None:
            _apply_nps_row(ws, r, col_map, nps_amount, nps_retro, nps_govt, result)
            result.nps_applied += 1

    # 매칭 통계
    raw_names = set()
    if nhis_data:
        raw_names.update(nhis_data.keys())
    if nps_member_data:
        raw_names.update(nps_member_data.keys())

    result.employees_matched = len(raw_names & wehago_names_seen)
    result.employees_unmatched = sorted(raw_names - wehago_names_seen)

    if result.employees_unmatched:
        result.warnings.append(
            f"매칭 실패 ({len(result.employees_unmatched)}명): "
            f"{', '.join(result.employees_unmatched[:5])}"
        )

    # 저장
    try:
        wb.save(upload_path)
    except Exception as e:
        result.warnings.append(f"저장 실패: {e}")
    finally:
        wb.close()

    logger.info(
        "Raw data 병합 완료: 매칭 %d명, NHIS %d명, NPS %d명, 경고 %d건",
        result.employees_matched, result.nhis_applied, result.nps_applied,
        len(result.warnings),
    )
    return result


def _apply_nhis_row(ws, row: int, col_map: dict, emp: NhisEmployeeData):
    """단일 행에 NHIS 데이터 덮어쓰기

    WEHAGO 컬럼 구조에 따라 3가지 패턴 지원:
      1. Simple:   건강보험료정산 (정산+연말+이자 통합)
      2. Standard: 건강보험정산   (정산+연말+이자 통합)
      3. Detailed: 연말정산(건강보험), 퇴직정산(건강보험) (분리)
    """
    # 건강보험 ← 산출_건강보험료
    col = col_map.get("건강보험")
    if col is not None:
        ws.cell(row, col + 1).value = emp.산출_건강보험료

    # 장기요양보험료 ← 산출_요양보험료
    col = col_map.get("장기요양보험료")
    if col is not None:
        ws.cell(row, col + 1).value = emp.산출_요양보험료

    # ── 정산 컬럼 처리: Detailed vs Simple/Standard ─────────────

    # Detailed 구조: 연말정산 컬럼이 별도로 있으면 분리 매핑
    has_연말정산_건강 = col_map.get("연말정산_건강") is not None
    has_연말정산_요양 = col_map.get("연말정산_요양") is not None

    if has_연말정산_건강:
        # 연말정산(건강보험) ← 연말정산_건강 + 이자_건강
        ws.cell(row, col_map["연말정산_건강"] + 1).value = (
            emp.연말정산_건강보험료 + emp.이자_건강보험료
        )
        # 건강보험정산 (있으면) ← 정산_건강 (퇴직 제외)
        col = col_map.get("건강보험정산")
        if col is not None:
            ws.cell(row, col + 1).value = emp.정산_건강보험료
        # 퇴직정산(건강보험) — 워크플로우 문서: "사용자가 수동으로 반영"
        # (자동 덮어쓰지 않음)
    else:
        # Simple/Standard 구조: 정산 컬럼에 모두 누적
        col = col_map.get("건강보험정산")
        if col is not None:
            total = emp.정산_건강보험료 + emp.연말정산_건강보험료 + emp.이자_건강보험료
            ws.cell(row, col + 1).value = total

    if has_연말정산_요양:
        # 연말정산(장기요양보험) ← 연말정산_요양 + 이자_요양
        ws.cell(row, col_map["연말정산_요양"] + 1).value = (
            emp.연말정산_요양보험료 + emp.이자_요양보험료
        )
        # 장기요양보험정산 (있으면) ← 정산_요양 (퇴직 제외)
        col = col_map.get("장기요양보험정산")
        if col is not None:
            ws.cell(row, col + 1).value = emp.정산_요양보험료
    else:
        # Simple/Standard 구조: 정산 컬럼에 모두 누적
        col = col_map.get("장기요양보험정산")
        if col is not None:
            total = emp.정산_요양보험료 + emp.연말정산_요양보험료 + emp.이자_요양보험료
            ws.cell(row, col + 1).value = total


def _apply_nps_row(ws, row: int, col_map: dict,
                   member_amount: int | None,
                   retro_amount: int | None,
                   govt_amount: int | None,
                   result: MergeResult):
    """단일 행에 NPS 데이터 덮어쓰기"""
    pension_col = col_map.get("국민연금")
    pension_settle_col = col_map.get("국민연금정산")

    # 기본: 국민연금 ← 근로자기여금
    base = member_amount if member_amount is not None else 0

    # 소급분 처리
    if retro_amount is not None:
        if pension_settle_col is not None:
            # 국민연금정산 컬럼이 있으면 소급분을 거기에
            ws.cell(row, pension_settle_col + 1).value = retro_amount
        else:
            # 국민연금정산 컬럼이 없으면 국민연금에 누적
            base += retro_amount

    # 국고지원금: 국민연금에서 차감 (절반을 음수로)
    if govt_amount is not None:
        base -= abs(govt_amount)

    if pension_col is not None:
        ws.cell(row, pension_col + 1).value = base
