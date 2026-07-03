"""건강보험(NHIS) PDF + 국민연금(NPS) Excel + 고용보험 XLS raw data 파싱 모듈

Phase 2/3/5에서 다운로드한 raw data 파일을 읽어 사원명 기준 dict로 변환.
Phase 6(WEHAGO 급여자료입력)에서 convert_for_upload로 전달됨.
"""
import logging
from dataclasses import dataclass, field

import openpyxl

logger = logging.getLogger(__name__)


# ═══════════════════════════════════════════════════════════════════════
# 공통 유틸
# ═══════════════════════════════════════════════════════════════════════

def _parse_int(val) -> int:
    """콤마 포함 숫자 문자열 → int. 빈값/None → 0."""
    if val is None:
        return 0
    s = str(val).strip().replace(",", "").replace(" ", "")
    if not s or s == "-":
        return 0
    try:
        return int(float(s))
    except (ValueError, TypeError):
        return 0


# ═══════════════════════════════════════════════════════════════════════
# NHIS (건강보험) PDF 파서
# ═══════════════════════════════════════════════════════════════════════

@dataclass
class NhisEmployeeData:
    """건강보험 가입자고지내역서에서 추출한 직원별 데이터"""
    name: str
    산출_건강보험료: int = 0
    산출_요양보험료: int = 0
    정산_건강보험료: int = 0       # 사유="퇴직" 제외
    정산_요양보험료: int = 0       # 사유="퇴직" 제외
    연말정산_건강보험료: int = 0
    연말정산_요양보험료: int = 0
    이자_건강보험료: int = 0
    이자_요양보험료: int = 0


def read_nhis_pdf(pdf_path: str) -> dict[str, NhisEmployeeData]:
    """NHIS 가입자고지내역서 PDF → {성명: NhisEmployeeData}

    PDF 테이블 구조 (직원당 2행: 건강 + 요양):
      idx0: 성명 (요양행은 None)
      idx4: 구분 ("건강" / "요양")
      idx5: 산출보험료
      idx7: 사유 (퇴직, 보수변경반환, 취득지연추가 등)
      idx9: 정산금액
      idx12: 연말정산보험료
      idx15: 환급금이자
    """
    import pdfplumber

    employees: dict[str, NhisEmployeeData] = {}

    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                tables = page.extract_tables()
                if not tables:
                    continue

                for table in tables:
                    _parse_nhis_table(table, employees)

    except Exception as e:
        logger.warning("NHIS PDF 파싱 실패 (%s): %s", pdf_path, e)
        return {}

    logger.info("NHIS PDF: %d명 파싱 완료 (%s)", len(employees), pdf_path)
    return employees


def _parse_nhis_table(table: list[list], employees: dict[str, NhisEmployeeData]):
    """NHIS PDF 테이블 행을 순회하며 직원별 데이터 누적"""
    current_name = None

    for row in table:
        if not row:
            continue

        # idx0: 성명 (요양행은 None → 직전 건강행의 이름 사용)
        name_val = row[0] if len(row) > 0 else None
        if name_val and str(name_val).strip():
            current_name = str(name_val).strip()
        if not current_name:
            continue

        # idx4: 구분 ("건강" / "요양")
        if len(row) <= 4 or not row[4]:
            continue
        구분 = str(row[4]).strip()
        if 구분 not in ("건강", "요양"):
            continue

        # 직원 데이터 확보
        if current_name not in employees:
            employees[current_name] = NhisEmployeeData(name=current_name)
        emp = employees[current_name]

        # idx7: 사유 — 퇴직 여부 확인
        사유 = str(row[7]).strip() if len(row) > 7 and row[7] else ""

        # idx5: 산출보험료
        산출 = _parse_int(row[5] if len(row) > 5 else None)

        # idx9: 정산금액 (퇴직 사유면 정산 컬럼에 넣지 않음)
        정산 = _parse_int(row[9] if len(row) > 9 else None)

        # idx12: 연말정산보험료
        연말 = _parse_int(row[12] if len(row) > 12 else None)

        # idx15: 환급금이자
        이자 = _parse_int(row[15] if len(row) > 15 else None)

        # 구분별 필드 매핑
        if 구분 == "건강":
            emp.산출_건강보험료 += 산출
            if "퇴직" not in 사유:
                emp.정산_건강보험료 += 정산
            emp.연말정산_건강보험료 += 연말
            emp.이자_건강보험료 += 이자
        else:  # 요양
            emp.산출_요양보험료 += 산출
            if "퇴직" not in 사유:
                emp.정산_요양보험료 += 정산
            emp.연말정산_요양보험료 += 연말
            emp.이자_요양보험료 += 이자


# ═══════════════════════════════════════════════════════════════════════
# NPS (국민연금) Excel 파서
# ═══════════════════════════════════════════════════════════════════════

def read_nps_member_excel(excel_path: str) -> dict[str, int]:
    """NPS 가입자내역 Excel → {성명: 근로자기여금}

    시트 구조:
      R3: 헤더 (순번, 성명, 주민등록번호, 기준소득월액, 연금보험료, 근로자기여금, 사용자부담금, ...)
      R5+: 데이터
    """
    result: dict[str, int] = {}

    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws = wb.active

        for r in range(5, ws.max_row + 1):
            name = ws.cell(r, 2).value  # C2: 성명
            if not name or not str(name).strip():
                continue
            name = str(name).strip()
            # C6: 근로자기여금
            contribution = _parse_int(ws.cell(r, 6).value)
            if name in result:
                result[name] += contribution
            else:
                result[name] = contribution

        wb.close()

    except Exception as e:
        logger.warning("NPS 가입자내역 파싱 실패 (%s): %s", excel_path, e)
        return {}

    logger.info("NPS 가입자내역: %d명 파싱 완료 (%s)", len(result), excel_path)
    return result


def read_nps_retro_excel(excel_path: str) -> dict[str, int]:
    """NPS 소급분내역 Excel → {성명: 본인기여금 합계}

    시트 구조:
      R3: 헤더 (순번, 성명, 주민등록번호, 기간, 월수, 발생금액, 사용자부담금, 본인기여금)
      R4+: 데이터 (동일 직원 다중 행 가능 → 합산)
    """
    result: dict[str, int] = {}

    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws = wb.active

        for r in range(4, ws.max_row + 1):
            name = ws.cell(r, 2).value  # C2: 성명
            if not name or not str(name).strip():
                continue
            name = str(name).strip()
            # C8: 본인기여금
            contribution = _parse_int(ws.cell(r, 8).value)
            if name in result:
                result[name] += contribution
            else:
                result[name] = contribution

        wb.close()

    except Exception as e:
        logger.warning("NPS 소급분내역 파싱 실패 (%s): %s", excel_path, e)
        return {}

    logger.info("NPS 소급분내역: %d명 파싱 완료 (%s)", len(result), excel_path)
    return result


def read_nps_govt_excel(excel_path: str) -> dict[str, int]:
    """NPS 국고지원내역 Excel → {성명: 국고지원금액(전액)/2}

    국고지원금액(근로자+사업장 합계, 전액)의 절반(=근로자 몫)을
    국민연금에서 (-)로 차감하기 위함.
    파일이 없거나 파싱 실패 시 빈 dict 반환.
    """
    result: dict[str, int] = {}

    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws = wb.active

        # 국고지원내역 구조 탐색
        support_col = None
        name_col = None
        # 국고지원금액을 사용자(사업장)/본인(근로자) 몫으로 나눈 sub-column은
        # 전액(국고지원금액) 컬럼 탐색에서 배제해야 함.
        # (이전엔 이 sub-column까지 매칭→마지막인 '본인기여금(X/2)'이 선택되었고,
        #  거기에 //2까지 더해져 차감액이 X/4로 과소 계산되었음)
        SPLIT_SUFFIXES = ("사용자부담금", "사업장", "본인기여금", "근로자")

        for r in range(1, min(ws.max_row + 1, 10)):
            for c in range(1, ws.max_column + 1):
                val = ws.cell(r, c).value
                if not val:
                    continue
                val_str = str(val).strip()
                if "성명" in val_str and name_col is None:
                    name_col = c
                # 국고지원금액(전액) 컬럼만 선택 — 분할 sub-column(사용자/본인 몫) 제외
                if ("지원금" in val_str or "전월분" in val_str) \
                        and not any(s in val_str for s in SPLIT_SUFFIXES):
                    support_col = c

        if not name_col or not support_col:
            logger.warning("NPS 국고지원내역: 필수 컬럼 미발견 (name_col=%s, support_col=%s)",
                           name_col, support_col)
            wb.close()
            return {}

        # 헤더 이후 데이터 행 읽기
        header_row = 3  # 일반적으로 R3가 헤더
        for r in range(header_row + 1, ws.max_row + 1):
            name = ws.cell(r, name_col).value
            if not name or not str(name).strip():
                continue
            name = str(name).strip()
            amount = _parse_int(ws.cell(r, support_col).value)
            if amount <= 0:
                continue  # 국고지원 미대상(0원) — 포함 시 국민연금 0 덮어쓰기 부작용 방지
            # 전액의 절반(=근로자 몫)을 국민연금에서 (-) 차감
            half_amount = amount // 2
            if name in result:
                result[name] += half_amount
            else:
                result[name] = half_amount

        wb.close()

    except Exception as e:
        logger.warning("NPS 국고지원내역 파싱 실패 (%s): %s", excel_path, e)
        return {}

    logger.info("NPS 국고지원내역: %d명 파싱 완료 (%s)", len(result), excel_path)
    return result


def read_nps_integrated_excel(
    excel_path: str,
) -> tuple[dict[str, int], dict[str, int], dict[str, int]]:
    """NPS 최종결정내역 탭 통합저장(전체표출) Excel → (member, retro, govt) 3개 dict.

    3개 탭(가입자/소급/국고) 개별 엑셀을 한 장으로 대체하는 단일 소스.
    컬럼을 헤더명으로 탐색(레이아웃 변경 내성):
      - member (당월분 본인기여금) = 근로자기여금   → 국민연금 기본값
      - retro  (소급분 본인기여금)                  → 국민연금정산/누적
      - govt   (국고지원금액 본인기여금, 이미 분할) → 국민연금에서 차감
    상실(퇴사) 행은 당월분 본인기여금이 공란이므로 member에서 제외되며,
    retro/govt는 0 초과인 경우만 포함(구 read_nps_retro/govt_excel 동작과 일치).

    Returns:
        (member_dict, retro_dict, govt_dict) — {성명: 금액}. 파싱 실패 시 빈 dict 3개.
    """
    member: dict[str, int] = {}
    retro: dict[str, int] = {}
    govt: dict[str, int] = {}

    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws = wb.active

        # 헤더 행 탐색 ("성명" 셀) + 3개 본인기여금 컬럼 식별.
        # "본인기여금"이 4컬럼(당월분/소급분/총부담금계/국고지원)에 공통 포함되므로
        # 접두사(당월분/소급분/국고지원)로 구분. 총부담금계(col19)는 의도적 배제.
        header_row = None
        name_col = member_col = retro_col = govt_col = None
        for r in range(1, min(ws.max_row + 1, 11)):
            for c in range(1, ws.max_column + 1):
                val = ws.cell(r, c).value
                if not val:
                    continue
                h = str(val).strip()
                if h == "성명" and name_col is None:
                    name_col = c
                    header_row = r
                if "본인기여금" in h:
                    if "당월분" in h and member_col is None:
                        member_col = c
                    elif "소급분" in h and retro_col is None:
                        retro_col = c
                    elif "국고지원" in h and govt_col is None:
                        govt_col = c

        if not header_row or not name_col:
            logger.warning(
                "NPS 통합엑셀: 헤더/성명 컬럼 미발견 (header_row=%s, name_col=%s)",
                header_row, name_col,
            )
            wb.close()
            return {}, {}, {}

        for r in range(header_row + 1, ws.max_row + 1):
            name = ws.cell(r, name_col).value
            if not name or not str(name).strip():
                continue
            name = str(name).strip()

            # member: 당월분 본인기여금 (공란=상실/퇴사 → 제외)
            if member_col is not None:
                mv = ws.cell(r, member_col).value
                if mv is not None and str(mv).strip() != "":
                    member[name] = member.get(name, 0) + _parse_int(mv)
            # retro: 소급분 본인기여금 (0 초과만)
            if retro_col is not None:
                rv = _parse_int(ws.cell(r, retro_col).value)
                if rv > 0:
                    retro[name] = retro.get(name, 0) + rv
            # govt: 국고지원 본인기여금 (0 초과만 — 이미 분할된 본인몫)
            if govt_col is not None:
                gv = _parse_int(ws.cell(r, govt_col).value)
                if gv > 0:
                    govt[name] = govt.get(name, 0) + gv

        wb.close()

    except Exception as e:
        logger.warning("NPS 통합엑셀 파싱 실패 (%s): %s", excel_path, e)
        return {}, {}, {}

    logger.info(
        "NPS 통합엑셀: member %d명, retro %d명, govt %d명 (%s)",
        len(member), len(retro), len(govt), excel_path,
    )
    return member, retro, govt


# ═══════════════════════════════════════════════════════════════════════
# 고용보험(근로복지공단) XLS 파서 — ClipReport 다운로드 파일
# ═══════════════════════════════════════════════════════════════════════

def read_employment_xls(xls_path: str) -> tuple[dict[str, int], dict[str, int]]:
    """고용보험료 지원금 정보 .xls 파일 파싱 (ClipReport 다운로드).

    엑셀 v3 규칙: 실업급여지원금(근로자) → 고용보험 (-), 환수금(근로자) → (+).
    본 함수는 근로자 몫만 추출하여 반환 (사업주 몫 제외).

    .xls 구조 (xlrd, OLE2 형식):
      - 시트 "Page 1" (지원금): col2=근로자명, col12=실업급여지원금(근로자)
      - 시트 "Page 2" (환수금): col1=근로자명, col7=실업급여환수금(근로자)

    Args:
        xls_path: 고용보험료지원금정보_{YYYYMM}.xls 경로

    Returns:
        (support_data, collect_data):
          support_data = {근로자명: 실업급여지원금(근로자)}
          collect_data = {근로자명: 실업급여환수금(근로자)}
    """
    import xlrd

    support: dict[str, int] = {}
    collect: dict[str, int] = {}

    try:
        wb = xlrd.open_workbook(xls_path)
    except Exception as e:
        logger.warning("고용보험 xls 열기 실패 (%s): %s", xls_path, e)
        return support, collect

    # Page 1 (지원금): col2=근로자명, col12=실업급여지원금(근로자)
    try:
        sh1 = wb.sheet_by_name("Page 1")
        for r in range(7, sh1.nrows):  # 행0-5 메타, 행6 헤더, 행7+ 데이터
            name = str(sh1.cell_value(r, 2)).strip()
            if not name:
                continue
            amount = _parse_int(sh1.cell_value(r, 12))
            if amount:
                support[name] = support.get(name, 0) + amount
    except Exception as e:
        logger.warning("고용보험 Page 1(지원금) 파싱 실패: %s", e)

    # Page 2 (환수금): col1=근로자명, col7=실업급여환수금(근로자)
    try:
        sh2 = wb.sheet_by_name("Page 2")
        for r in range(7, sh2.nrows):
            name = str(sh2.cell_value(r, 1)).strip()
            if not name:
                continue
            amount = _parse_int(sh2.cell_value(r, 7))
            if amount:
                collect[name] = collect.get(name, 0) + amount
    except Exception as e:
        logger.warning("고용보험 Page 2(환수금) 파싱 실패: %s", e)

    logger.info(
        "고용보험 xls: 지원금 %d명, 환수금 %d명 (%s)",
        len(support), len(collect), xls_path,
    )
    return support, collect
