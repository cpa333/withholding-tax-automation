"""read_nps_integrated_excel 파서 단위테스트.

NPS 최종결정내역 통합엑셀 파서의 헤더 탐색/빈값/합산/오프셋/엣지케이스 검증.
라이브 데이터 없이 openpyxl 로 가짜 엑셀을 생성해 순수 단위테스트.
(PHASE 3 통합엑셀 단일화 — col10/col16/col24 → member/retro/govt)
"""
from openpyxl import Workbook

from src.utils.raw_data_reader import read_nps_integrated_excel


# 실제 NPS 통합엑셀 헤더(R3) — col10/16/24 가 본인기여금.
HEADERS = [
    "고지년월", "사업장관리번호", "성명", "주민등록번호",
    "취득일(공란은 당월 이전 취득중인 자)", "상실일(공란은 취득중인 자)",
    "당월분_기준소득월액(원)", "당월분_월보험료(원)",
    "당월분_(사용자부담금)(원)", "당월분_(본인기여금(원)",
    "소급분_변동사유", "소급분_해당기간", "소급분_월수",
    "소급분_전월이전보험료(원)", "소급분_(사용자부담금)(원)",
    "소급분_(본인기여금)(원)", "이자분(0원)",
    "총부담금계_(사용자부담금)(원)", "총부담금계_(본인기여금)(원)",
    "자격변동 신고사항", "자격변동 신고사항_대상자",
    "국고지원금액(원)", "국고지원금액_사용자부담금(원)",
    "국고지원금액_본인기여금(원)", "취득월 납부여부",
]


def _row(name, member=None, retro=None, govt=None):
    """HEADERS 기준 25-컬럼 행. None=미설정(공란)."""
    row = [None] * 25
    row[0] = "2026-06"
    row[1] = "11286016310"
    row[2] = name
    row[3] = "900101-1000000"
    if member is not None:
        row[9] = member    # col10 당월분 본인기여금
    if retro is not None:
        row[15] = retro    # col16 소급분 본인기여금
    if govt is not None:
        row[23] = govt     # col24 국고 본인기여금
    return row


def _write(tmp_path, name, header_row, data_rows, title_rows=0):
    """가짜 엑셀 생성. title_rows>0 면 헤더 전 잡동사니 행 삽입(헤더 offset)."""
    path = tmp_path / name
    wb = Workbook()
    ws = wb.active
    ws.title = "sheet"
    r = 1
    for _ in range(title_rows):
        ws.cell(r, 1).value = f"잡동사니{r}"
        r += 1
    if header_row:
        for c, val in enumerate(header_row, start=1):
            ws.cell(r, c).value = val
        r += 1
    for dr in data_rows:
        for c, val in enumerate(dr, start=1):
            ws.cell(r, c).value = val
        r += 1
    wb.save(path)
    wb.close()
    return str(path)


def test_normal_sample(tmp_path):
    """정상: 활성 가입자 member 값, 상실자(col10 공란) 제외."""
    path = _write(tmp_path, "normal.xlsx", HEADERS, [
        _row("홍길동", member=154370),
        _row("김철수", member=204150, retro=10000, govt=5000),
        _row("퇴사자", member=None),  # 상실
    ])
    m, r, g = read_nps_integrated_excel(path)
    assert m == {"홍길동": 154370, "김철수": 204150}
    assert r == {"김철수": 10000}
    assert g == {"김철수": 5000}
    assert "퇴사자" not in m


def test_empty_workbook(tmp_path):
    path = str(tmp_path / "empty.xlsx")
    wb = Workbook()
    wb.save(path)
    wb.close()
    assert read_nps_integrated_excel(path) == ({}, {}, {})


def test_header_only_no_data(tmp_path):
    path = _write(tmp_path, "hdr.xlsx", HEADERS, [])
    assert read_nps_integrated_excel(path) == ({}, {}, {})


def test_missing_contribution_columns(tmp_path):
    """본인기여금 헤더 자체가 없으면 member/retro/govt 모두 빈(크래시 X)."""
    headers = ["고지년월", "사업장관리번호", "성명", "주민등록번호", "기준소득월액"]
    rows = [["2026-06", "11286016310", "홍길동", "900101-1000000", 3250000]]
    path = _write(tmp_path, "nocontrib.xlsx", headers, rows)
    assert read_nps_integrated_excel(path) == ({}, {}, {})


def test_missing_name_column(tmp_path):
    headers = ["고지년월", "당월분_(본인기여금(원)"]
    rows = [["2026-06", 154370]]
    path = _write(tmp_path, "noname.xlsx", headers, rows)
    assert read_nps_integrated_excel(path) == ({}, {}, {})


def test_all_loss_only(tmp_path):
    """모든 col10 공란(상실-only) → member 빈."""
    path = _write(tmp_path, "loss.xlsx", HEADERS, [
        _row("퇴사자A", member=None),
        _row("퇴사자B", member=None),
    ])
    assert read_nps_integrated_excel(path) == ({}, {}, {})


def test_retro_govt_positive_only(tmp_path):
    """retro/govt는 0 초과만 포함. 0인 사람은 미포함."""
    path = _write(tmp_path, "rg.xlsx", HEADERS, [
        _row("정상", member=100000, retro=0, govt=0),
        _row("소급", member=50000, retro=30000, govt=0),
        _row("국고", member=40000, retro=0, govt=20000),
    ])
    m, r, g = read_nps_integrated_excel(path)
    assert m == {"정상": 100000, "소급": 50000, "국고": 40000}
    assert r == {"소급": 30000}
    assert g == {"국고": 20000}


def test_duplicate_name_merge(tmp_path):
    """동일 성명 2행 → 각 dict 합산."""
    path = _write(tmp_path, "dup.xlsx", HEADERS, [
        _row("동명", member=100000, retro=10000, govt=5000),
        _row("동명", member=50000, retro=20000, govt=3000),
    ])
    m, r, g = read_nps_integrated_excel(path)
    assert m == {"동명": 150000}
    assert r == {"동명": 30000}
    assert g == {"동명": 8000}


def test_header_offset(tmp_path):
    """헤더가 5번째 행에 있어도 정상 탐지."""
    path = _write(tmp_path, "offset.xlsx", HEADERS,
                  [_row("오프셋", member=111111, retro=2222, govt=333)],
                  title_rows=4)
    m, r, g = read_nps_integrated_excel(path)
    assert m == {"오프셋": 111111}
    assert r == {"오프셋": 2222}
    assert g == {"오프셋": 333}


def test_comma_and_whitespace_parsing(tmp_path):
    """콤마 포함 문자열 → 정수, whitespace 공란 → member 제외, 이름 공백 트림."""
    wb = Workbook()
    ws = wb.active
    for c, val in enumerate(HEADERS, start=1):
        ws.cell(1, c).value = val
    ws.cell(2, 3).value = "콤마맨"
    ws.cell(2, 10).value = "154,370"   # → 154370
    ws.cell(3, 3).value = "공백맨"
    ws.cell(3, 10).value = "   "       # whitespace → 공란 → member 제외
    ws.cell(4, 3).value = "  트림  "
    ws.cell(4, 10).value = 777
    path = str(tmp_path / "mixed.xlsx")
    wb.save(path)
    wb.close()
    m, _, _ = read_nps_integrated_excel(path)
    assert m.get("콤마맨") == 154370
    assert "공백맨" not in m
    assert m.get("트림") == 777


def test_missing_file_returns_empty():
    import os
    missing = os.path.join(os.path.dirname(__file__), "nope_definitely_missing.xlsx")
    assert not os.path.exists(missing)
    assert read_nps_integrated_excel(missing) == ({}, {}, {})


def test_float_cells(tmp_path):
    """openpyxl 이 숫자 셀을 float(154370.0)로 반환하는 경우 — _parse_int 의
    int(float(s)) 경로. 실제 NPS 통합엑셀은 float 셀을 쓰므로 핵심 경로.
    이 경로가 깨지면(int(s) 회귀) '154370.0' → ValueError → 0 로 모든 값 추락."""
    wb = Workbook()
    ws = wb.active
    for c, val in enumerate(HEADERS, start=1):
        ws.cell(1, c).value = val
    ws.cell(2, 3).value = "실수맨"
    ws.cell(2, 10).value = 154370.0     # float — openpyxl 실제 반환 타입
    ws.cell(2, 16).value = 10000.5      # float 소수 → int truncation 10000
    ws.cell(2, 24).value = 5000.0
    ws.cell(3, 3).value = "문자열실수"
    ws.cell(3, 10).value = "204150.0"   # 문자열 '204150.0' 도 int(float) 처리
    path = str(tmp_path / "float.xlsx")
    wb.save(path)
    wb.close()
    m, r, g = read_nps_integrated_excel(path)
    assert m == {"실수맨": 154370, "문자열실수": 204150}
    assert r == {"실수맨": 10000}   # 10000.5 → truncation
    assert g == {"실수맨": 5000}


def test_col19_total_excluded(tmp_path):
    """총부담금계(col19)는 '본인기여금' 포함이나 접두사(당월/소급/국고) 없어
    매칭에서 배제됨 → col10 만 member 로 사용(col19 값은 무시). 동작보존 핵심."""
    wb = Workbook()
    ws = wb.active
    for c, val in enumerate(HEADERS, start=1):
        ws.cell(1, c).value = val
    ws.cell(2, 3).value = "테스트"
    ws.cell(2, 10).value = 100000      # col10 당월분 본인기여금
    ws.cell(2, 19).value = 999999      # col19 총부담금계 (무시되어야)
    path = str(tmp_path / "col19.xlsx")
    wb.save(path)
    wb.close()
    m, _, _ = read_nps_integrated_excel(path)
    assert m == {"테스트": 100000}
