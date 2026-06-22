"""NPS→WEHAGO 병합 동작보존 회귀 테스트.

통합엑셀 경로(read_nps_integrated_excel)와 구 3파일 경로(read_nps_member/
retro/govt_excel)가 동일한 underlying 값에서 apply_raw_data 를 거쳐
WEHAGO 업로드 엑셀의 국민연금/국민연금정산 셀에 같은 값을 쓰는지 검증.
member / retro-only / govt(차감) 케이스를 합성 데이터로 커버.
"""
from openpyxl import Workbook, load_workbook

from src.utils.raw_data_reader import (
    read_nps_integrated_excel, read_nps_member_excel,
    read_nps_retro_excel, read_nps_govt_excel,
)
from src.utils.data_merger import apply_raw_data


# 통합엑셀 헤더 — col3 성명 / col10 당월분 본인기여금 / col16 소급분 / col24 국고 본인기여금
HEADERS = [
    "고지년월", "사업장관리번호", "성명", "주민등록번호", "취득일", "상실일",
    "당월분_기준소득월액(원)", "당월분_월보험료(원)",
    "당월분_(사용자부담금)(원)", "당월분_(본인기여금(원)",
    "소급분_변동사유", "소급분_해당기간", "소급분_월수",
    "소급분_전월이전보험료(원)", "소급분_(사용자부담금)(원)",
    "소급분_(본인기여금)(원)", "이자분(0원)",
    "총부담금계_(사용자부담금)(원)", "총부담금계_(본인기여금)(원)",
    "자격변동 신고사항", "자격변동 신고사항_대상자",
    "국고지원금액(원)", "국고지원금액_사용자부담금(원)",
    "국고지원금액_본인기여금(원)", "취득월 납부여부",
]


def _int_row(name, member=None, retro=None, govt=None):
    row = [None] * 25
    row[2] = name
    if member is not None:
        row[9] = member    # col10
    if retro is not None:
        row[15] = retro    # col16
    if govt is not None:
        row[23] = govt     # col24
    return row


def _write_integrated(path, rows):
    wb = Workbook(); ws = wb.active
    for c, val in enumerate(HEADERS, start=1):
        ws.cell(1, c).value = val
    r = 2
    for row in rows:
        for c, val in enumerate(row, start=1):
            ws.cell(r, c).value = val
        r += 1
    wb.save(path); wb.close()


def _write_member_file(path, d):
    """구 가입자내역: R3 헤더, C2 성명, C6 근로자기여금, R5+ 데이터."""
    wb = Workbook(); ws = wb.active
    ws.cell(3, 2).value = "성명"; ws.cell(3, 6).value = "근로자기여금"
    r = 5
    for n, v in d.items():
        ws.cell(r, 2).value = n; ws.cell(r, 6).value = v; r += 1
    wb.save(path); wb.close()


def _write_retro_file(path, d):
    """구 소급분내역: R3 헤더, C2 성명, C8 본인기여금, R4+ 데이터."""
    wb = Workbook(); ws = wb.active
    ws.cell(3, 2).value = "성명"; ws.cell(3, 8).value = "본인기여금"
    r = 4
    for n, v in d.items():
        ws.cell(r, 2).value = n; ws.cell(r, 8).value = v; r += 1
    wb.save(path); wb.close()


def _write_govt_file(path, d):
    """구 국고지원내역: 전액 컬럼(//2 되므로 본인몫×2)."""
    wb = Workbook(); ws = wb.active
    ws.cell(3, 2).value = "성명"; ws.cell(3, 3).value = "국고지원금액(원)"
    r = 4
    for n, v in d.items():
        ws.cell(r, 2).value = n; ws.cell(r, 3).value = v * 2; r += 1
    wb.save(path); wb.close()


def _write_wehago(path, names):
    wb = Workbook(); ws = wb.active
    ws.cell(1, 1).value = "사원명"
    ws.cell(1, 2).value = "국민연금"
    ws.cell(1, 3).value = "국민연금정산"
    for i, n in enumerate(names, start=2):
        ws.cell(i, 1).value = n
        ws.cell(i, 2).value = 0
        ws.cell(i, 3).value = 0
    wb.save(path); wb.close()


def _read_cells(path):
    wb = load_workbook(path); ws = wb.active
    out = {}
    for r in range(2, ws.max_row + 1):
        n = ws.cell(r, 1).value
        if n:
            out[str(n).strip()] = (ws.cell(r, 2).value, ws.cell(r, 3).value)
    wb.close(); return out


def test_integrated_path_matches_three_file_path(tmp_path):
    """통합엑셀 경로와 구 3파일 경로가 같은 WEHAGO 셀 값을 낸다 (동작보존).

    시나리오:
      홍길동: member 100000 + retro 20000          → 국민연금 100000, 정산 20000
      이순신: retro-only 50000                       → 국민연금 0(미건드림), 정산 50000
      김국고: member 80000 + govt(본인몫) 30000      → 국민연금 50000(차감), 정산 0
    """
    rows = [
        _int_row("홍길동", member=100000, retro=20000),
        _int_row("이순신", retro=50000),           # member 공란 → member dict 제외
        _int_row("김국고", member=80000, govt=30000),
    ]
    int_path = str(tmp_path / "integrated.xlsx")
    _write_integrated(int_path, rows)
    m, r, g = read_nps_integrated_excel(int_path)
    assert m == {"홍길동": 100000, "김국고": 80000}
    assert r == {"홍길동": 20000, "이순신": 50000}
    assert g == {"김국고": 30000}

    # 구 3파일: 동일 underlying 값 합성
    mf = str(tmp_path / "m.xlsx"); _write_member_file(mf, m)
    rf = str(tmp_path / "r.xlsx"); _write_retro_file(rf, r)
    gf = str(tmp_path / "g.xlsx"); _write_govt_file(gf, g)
    m_old = read_nps_member_excel(mf)
    r_old = read_nps_retro_excel(rf)
    g_old = read_nps_govt_excel(gf)
    assert m_old == m
    assert r_old == r
    assert g_old == g

    names = ["홍길동", "이순신", "김국고"]
    up_int = str(tmp_path / "wehago_int.xlsx"); _write_wehago(up_int, names)
    up_old = str(tmp_path / "wehago_old.xlsx"); _write_wehago(up_old, names)
    apply_raw_data(up_int, nps_member_data=m, nps_retro_data=r, nps_govt_data=g)
    apply_raw_data(up_old, nps_member_data=m_old, nps_retro_data=r_old, nps_govt_data=g_old)

    cells_int = _read_cells(up_int)
    cells_old = _read_cells(up_old)
    # 동작보존: 두 경로 셀 완전 일치
    assert cells_int == cells_old
    # 기대 수치 (member→국민연금, retro→정산, govt→국민연금 차감)
    assert cells_int["홍길동"] == (100000, 20000)
    assert cells_int["이순신"] == (0, 50000)
    assert cells_int["김국고"] == (50000, 0)
